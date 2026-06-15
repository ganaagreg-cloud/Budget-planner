"""
Ultimate Budget Planner generator.

Builds "Ultimate Budget Planner.xlsx" - a 14-sheet personal finance workbook
designed to work identically in Excel and Google Sheets.

CELL MAP (cross-sheet reference contract - keep in sync when editing layouts):
  START HERE:        USER_NAME=B5  CURRENCY=B6  CURRENT_MONTH=B7
                      MONTHLY_INCOME=B8  SAVINGS_GOAL_PCT=B9
  Income Tracker:    data rows 4-103 | D=Amount  H=Month("MMMM YYYY")
  Expense Tracker:   data rows 4-503 | C=Category D=Subcategory E=Amount
                      G=Need/Want J=Month("MMMM YYYY")
  Monthly Budget:    data rows 4-103 | B=Category E=Budgeted F=Actual G=Diff H=%Used
  Savings Goals:     6 goal blocks, 5 rows each starting at row 4 (rows 4-33)
                      row+0: A=Name (merged A:D)  F=Target (merged F:H)
                      row+1: B=Saved D=Monthly Contribution F=Target Date H=Remaining
                      row+2: A=Progress bar (merged A:H)
                      row+3: A=Projection text (merged A:H)
                      row+4: spacer
                      row 36: totals (B=Saved D=Target F=Remaining H=%)
  Debt Tracker:      data rows 4-7 (4 debts) | D=Current Balance L=Progress%
                      A=Name B=Type C=Starting Balance E=Rate F=Min Payment
                      G=Extra Payment H=Total Payment I=Monthly Interest
                      J=Principal Paid K=Est. Payoff Date M=Progress bar
                      N=Payoff date value (hidden helper)
  Sinking Funds:     data rows 4-15 (12 funds) | C=Annual Target D=Current Balance
                      E=Suggested Monthly F=Remaining G=% Funded H=Progress bar
  Subscriptions:     data rows 4-15 (12 subs)  | C=Monthly Cost E=Annual Cost H=Active?
  Bills Calendar:    data rows 4-18 (15 bills) | C=Due Day D=Amount G=Paid? L=Status
"""

import os
import sys
from datetime import date, timedelta

if hasattr(sys.stdout, "reconfigure"):
    sys.stdout.reconfigure(encoding="utf-8")

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Border, Side, Alignment, Protection
from openpyxl.utils import get_column_letter, range_boundaries
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.formatting.rule import CellIsRule, FormulaRule, Rule
from openpyxl.styles.differential import DifferentialStyle
from openpyxl.workbook.defined_name import DefinedName
from openpyxl.comments import Comment

# ============================================================
# SECTION 1: CONSTANTS
# ============================================================
PRIMARY_GREEN = "C9B8B8"   # dusty rose/mauve — title header fills
LIGHT_GREEN = "B2D8D8"    # mint green — positive value fills, savings
PALE_GREEN = "D4EDEA"     # light mint — alternating rows
WHITE = "FFFFFF"
LIGHT_GRAY = "FDF6F0"     # warm cream — background / stat bars
MEDIUM_GRAY = "E9ECEF"
DARK_GRAY = "4A3728"      # warm dark brown — secondary text
ACCENT_BLUE = "E8D5D5"    # soft pink — section header fills
SOFT_RED = "D4756B"       # warm red — over budget
LIGHT_RED = "FFE5E5"
GOLD = "C9A84C"           # gold accent — KPI highlights, icons
DARK_TEXT = "4A3728"      # warm dark brown — primary text
POSITIVE_TEXT = "1A7272"  # dark teal — positive value / progress bar text
YELLOW_INPUT = "FFF3CD"
LAVENDER = "F3E8FF"
LIGHT_BLUE = "D6EAF8"
LIGHT_ORANGE = "FDE6D2"

FONT_NAME = "Calibri"
CURRENCY_FMT = "#,##0.00"
PERCENT_FMT = "0.0%"
DATE_FMT = "YYYY-MM-DD"
INT_FMT = "#,##0"

THIN_SIDE = Side(style="thin", color=MEDIUM_GRAY)
MEDIUM_SIDE = Side(style="medium", color=DARK_GRAY)

BACK_TO_DASHBOARD = "← Back to Dashboard"

CATEGORIES_LIST = [
    "Housing", "Food & Dining", "Transportation", "Entertainment",
    "Shopping", "Healthcare", "Utilities", "Education", "Personal Care",
    "Savings", "Investments", "Debt Payment", "Kids", "Pets", "Travel",
    "Subscriptions", "Insurance", "Other",
]
SUBCATEGORIES_LIST = [
    "Rent", "Mortgage", "Home Insurance", "Repairs", "HOA",
    "Groceries", "Restaurants", "Coffee", "Delivery",
    "Gas", "Car Insurance", "Maintenance", "Parking", "Public Transit",
    "Streaming", "Movies", "Games", "Sports", "Hobbies",
    "Clothing", "Electronics", "Gifts",
    "Doctor", "Dental", "Pharmacy", "Vision",
    "Electricity", "Water", "Internet", "Phone", "Gas Utility",
    "Tuition", "Books", "Courses",
    "Haircuts", "Gym", "Spa",
    "Emergency Fund", "General Savings",
    "Stocks", "Retirement",
    "Credit Card", "Loan Payment",
    "School Supplies", "Childcare", "Activities",
    "Pet Food", "Vet", "Grooming",
    "Flights", "Hotels", "Trip Activities",
    "Other",
]
PAYMENT_METHODS_LIST = [
    "Cash", "Debit Card", "Credit Card", "Bank Transfer",
    "PayPal", "Venmo", "Check", "Other",
]
NEED_WANT_LIST = ["Need", "Want", "Investment", "Savings"]
YES_NO_LIST = ["Yes", "No"]
INCOME_SOURCES_LIST = [
    "Salary", "Freelance/Contract", "Business Revenue", "Rental Income",
    "Dividends", "Interest", "Refund", "Gift", "Side Hustle", "Other",
]
BILLING_CYCLES_LIST = ["Monthly", "Annual", "Quarterly", "Weekly"]
DEBT_TYPES_LIST = [
    "Credit Card", "Car Loan", "Student Loan", "Personal Loan",
    "Medical Debt", "Mortgage", "Home Equity", "Business Loan", "Other",
]
PAID_STATUS_LIST = ["Yes", "No", "Pending"]
SUB_STATUS_LIST = ["Active", "Cancelled", "Paused"]
CATEGORY_TYPE_LIST = ["Active", "Passive"]

# ============================================================
# SECTION 2: STYLE HELPER FUNCTIONS
# ============================================================

def style_title(cell, text=None):
    """Dusty rose header used for sheet title banners."""
    if text is not None:
        cell.value = text
    cell.font = Font(name=FONT_NAME, size=20, bold=True, color=DARK_TEXT)
    cell.fill = PatternFill("solid", fgColor=PRIMARY_GREEN)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_section_header(cell, text=None):
    """Soft pink section sub-header."""
    if text is not None:
        cell.value = text
    cell.font = Font(name=FONT_NAME, size=13, bold=False, color=DARK_TEXT)
    cell.fill = PatternFill("solid", fgColor=ACCENT_BLUE)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_col_header(cell, text=None):
    """Pale green column label with border."""
    if text is not None:
        cell.value = text
    cell.font = Font(name=FONT_NAME, size=11, bold=True, color=DARK_TEXT)
    cell.fill = PatternFill("solid", fgColor=PALE_GREEN)
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    apply_border(cell)


def style_kpi_label(cell, text=None):
    if text is not None:
        cell.value = text
    cell.font = Font(name=FONT_NAME, size=10, color=DARK_GRAY)
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_kpi(cell, value=None, color=DARK_TEXT, currency=False, percent=False):
    """Large bold KPI number."""
    if value is not None:
        cell.value = value
    cell.font = Font(name=FONT_NAME, size=18, bold=True, color=color)
    if currency:
        cell.number_format = CURRENCY_FMT
    elif percent:
        cell.number_format = PERCENT_FMT
    cell.alignment = Alignment(horizontal="left", vertical="center")


def style_positive(cell):
    cell.fill = PatternFill("solid", fgColor=LIGHT_GREEN)
    cell.font = Font(name=FONT_NAME, color=POSITIVE_TEXT, bold=True)


def style_negative(cell):
    cell.fill = PatternFill("solid", fgColor=LIGHT_RED)
    cell.font = Font(name=FONT_NAME, color=SOFT_RED, bold=True)


def style_currency(cell):
    cell.number_format = CURRENCY_FMT


def style_percent(cell):
    cell.number_format = PERCENT_FMT


def style_date_cell(cell):
    cell.number_format = DATE_FMT


def style_input(cell, value=None):
    """Yellow, unlocked input cell."""
    if value is not None:
        cell.value = value
    cell.fill = PatternFill("solid", fgColor=YELLOW_INPUT)
    cell.protection = Protection(locked=False)
    apply_border(cell)


def apply_border(cell, style="thin"):
    side = THIN_SIDE if style == "thin" else MEDIUM_SIDE
    cell.border = Border(left=side, right=side, top=side, bottom=side)


def apply_table_borders(ws, cell_range):
    """Thick outer border, thin inner borders over a range like 'A1:H10'."""
    min_col, min_row, max_col, max_row = range_boundaries(cell_range)
    for row in range(min_row, max_row + 1):
        for col in range(min_col, max_col + 1):
            cell = ws.cell(row=row, column=col)
            left = MEDIUM_SIDE if col == min_col else THIN_SIDE
            right = MEDIUM_SIDE if col == max_col else THIN_SIDE
            top = MEDIUM_SIDE if row == min_row else THIN_SIDE
            bottom = MEDIUM_SIDE if row == max_row else THIN_SIDE
            cell.border = Border(left=left, right=right, top=top, bottom=bottom)


def alternate_row_colors(ws, min_row, max_row, min_col, max_col):
    """Alternate WHITE / PALE_GREEN row backgrounds over a data range."""
    for i, row in enumerate(range(min_row, max_row + 1)):
        if i % 2 == 1:
            for col in range(min_col, max_col + 1):
                cell = ws.cell(row=row, column=col)
                cell.fill = PatternFill("solid", fgColor=PALE_GREEN)


def set_col_widths(ws, widths):
    """widths: dict like {'A': 12, 'B': 18, ...}"""
    for col, width in widths.items():
        ws.column_dimensions[col].width = width


def progress_bar_formula(pct_ref):
    """REPT-based 20-segment progress bar with trailing percentage."""
    return (
        f'=REPT("█",ROUND({pct_ref}*20,0))&'
        f'REPT("░",20-ROUND({pct_ref}*20,0))&" "&TEXT({pct_ref},"0%")'
    )


def cf_diff_green_red(ws, cell_range):
    """Green fill for positive values, red fill for negative values."""
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=LIGHT_GREEN),
                   font=Font(name=FONT_NAME, color=POSITIVE_TEXT)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0"],
                   fill=PatternFill("solid", fgColor=LIGHT_RED),
                   font=Font(name=FONT_NAME, color=SOFT_RED)),
    )


def cf_percent_used(ws, cell_range):
    """Red if >100%, gold if 80-100%, green if <80%."""
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="greaterThan", formula=["1"],
                   fill=PatternFill("solid", fgColor=LIGHT_RED)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="between", formula=["0.8", "1"],
                   fill=PatternFill("solid", fgColor=GOLD)),
    )
    ws.conditional_formatting.add(
        cell_range,
        CellIsRule(operator="lessThan", formula=["0.8"],
                   fill=PatternFill("solid", fgColor=LIGHT_GREEN)),
    )


def progress_bar_expr(pct_expr):
    """REPT-based 20-segment progress bar built from an arbitrary percent expression."""
    return (
        f'=REPT("█",ROUND(({pct_expr})*20,0))&'
        f'REPT("░",20-ROUND(({pct_expr})*20,0))&" "&TEXT({pct_expr},"0%")'
    )


def add_dropdown(ws, cell_range, named_range, allow_blank=True):
    dv = DataValidation(type="list", formula1=f"={named_range}", allow_blank=allow_blank)
    ws.add_data_validation(dv)
    dv.add(cell_range)
    return dv


def add_back_link(ws):
    """Adds a small note on A1 pointing back to the Dashboard."""
    ws["A1"].comment = Comment(BACK_TO_DASHBOARD, "Budget Planner")


def finalize_sheet(ws, protect=True):
    """Common finishing touches: landscape page, header/footer, protection."""
    ws.sheet_view.showGridLines = False
    ws.page_setup.orientation = "landscape"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr = None
    ws.oddHeader.center.text = f"Ultimate Budget Planner | {ws.title} | Page &P"
    ws.oddFooter.center.text = "Generated with Budget Planner Pro"
    if protect:
        ws.protection.sheet = True
        ws.protection.formatCells = False
        ws.protection.formatColumns = False
        ws.protection.formatRows = False
        ws.protection.autoFilter = False
        ws.protection.sort = False


# ============================================================
# SECTION 3: SHEET BUILDER FUNCTIONS
# ============================================================

def build_lists(wb):
    print("✅ Building LISTS...")
    ws = wb.create_sheet("LISTS")

    lists = {
        "A": ("Category", CATEGORIES_LIST),
        "B": ("Subcategory", SUBCATEGORIES_LIST),
        "C": ("Payment_Method", PAYMENT_METHODS_LIST),
        "D": ("Need_Want", NEED_WANT_LIST),
        "E": ("Yes_No", YES_NO_LIST),
        "F": ("Income_Source", INCOME_SOURCES_LIST),
        "G": ("Billing_Cycle", BILLING_CYCLES_LIST),
        "H": ("Debt_Type", DEBT_TYPES_LIST),
        "I": ("Paid_Status", PAID_STATUS_LIST),
        "J": ("Sub_Status", SUB_STATUS_LIST),
        "K": ("Category_Type", CATEGORY_TYPE_LIST),
    }

    name_map = {
        "A": "CATEGORIES", "B": "SUBCATEGORIES", "C": "PAYMENT_METHODS",
        "D": "NEED_WANT", "E": "YES_NO", "F": "INCOME_SOURCES",
        "G": "BILLING_CYCLES", "H": "DEBT_TYPES", "I": "PAID_STATUS",
        "J": "SUB_STATUS", "K": "CATEGORY_TYPE",
    }

    for col, (header, items) in lists.items():
        ws[f"{col}1"] = header
        style_col_header(ws[f"{col}1"])
        for i, item in enumerate(items, start=2):
            ws[f"{col}{i}"] = item
        ws.column_dimensions[col].width = 20
        last_row = len(items) + 1
        ref = f"LISTS!${col}$2:${col}${last_row}"
        wb.defined_names.add(DefinedName(name_map[col], attr_text=ref))

    ws.sheet_state = "hidden"
    return ws


def build_start_here(wb):
    print("✅ Building START HERE...")
    ws = wb.create_sheet("START HERE")
    set_col_widths(ws, {"A": 28, "B": 22, "C": 45, "D": 14, "E": 14, "F": 14, "G": 14, "H": 14})

    ws.merge_cells("A1:H1")
    style_title(ws["A1"], "\U0001F49A Ultimate Budget Planner")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = "Your complete financial management system"
    ws["A2"].font = Font(name=FONT_NAME, size=12, italic=True, color=DARK_GRAY)

    # Setup section
    ws.merge_cells("A4:H4")
    style_section_header(ws["A4"], "⚙️ SETUP — Enter Your Info Here (Only Edit Yellow Cells)")
    ws.row_dimensions[4].height = 32

    setup_rows = [
        (5, "Your Name", "Alex Johnson", "USER_NAME"),
        (6, "Currency Symbol", "$", "CURRENCY"),
        (7, "Current Month", '=TEXT(TODAY(),"MMMM YYYY")', "CURRENT_MONTH"),
        (8, "Monthly Income", 5500, "MONTHLY_INCOME"),
        (9, "Monthly Savings Goal %", 0.2, "SAVINGS_GOAL_PCT"),
    ]
    for row, label, example, name in setup_rows:
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name=FONT_NAME, bold=True, color=DARK_TEXT)
        style_input(ws[f"B{row}"], example)
        if name == "MONTHLY_INCOME":
            style_currency(ws[f"B{row}"])
        if name == "SAVINGS_GOAL_PCT":
            style_percent(ws[f"B{row}"])
        ws[f"C{row}"] = f"Example: {example}" if not str(example).startswith("=") else "Auto-fills from today's date"
        ws[f"C{row}"].font = Font(name=FONT_NAME, size=9, italic=True, color=DARK_GRAY)
        wb.defined_names.add(DefinedName(name, attr_text=f"'START HERE'!$B${row}"))

    # How to use section
    ws.merge_cells("A11:H11")
    style_section_header(ws["A11"], "\U0001F4D6 HOW TO USE THIS PLANNER")
    ws.row_dimensions[11].height = 32

    steps = [
        ("Expense Tracker", "Log every expense daily as it happens."),
        ("Monthly Budget", "Plan your spending limits for each category."),
        ("Dashboard", "Check your progress weekly at a glance."),
        ("Savings Goals", "Set and track targets for what matters most."),
        ("Bills Calendar", "Track every bill so you never miss a payment."),
        ("Income Tracker", "Log all income sources as they arrive."),
        ("Debt Tracker", "Plan your debt payoff with avalanche or snowball."),
        ("Annual Overview", "Review trends at the end of each month."),
    ]
    for i, (title, desc) in enumerate(steps):
        row = 12 + i
        ws[f"A{row}"] = f"{i + 1}."
        ws[f"A{row}"].font = Font(name=FONT_NAME, bold=True, color=GOLD, size=12)
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"] = title
        ws[f"B{row}"].font = Font(name=FONT_NAME, bold=True, color=DARK_TEXT)
        ws.merge_cells(f"D{row}:H{row}")
        ws[f"D{row}"] = desc
        ws[f"D{row}"].font = Font(name=FONT_NAME, color=DARK_GRAY)

    # Sheet guide
    guide_row = 21
    ws.merge_cells(f"A{guide_row}:H{guide_row}")
    style_section_header(ws[f"A{guide_row}"], "\U0001F5FA️ SHEET GUIDE")
    ws.row_dimensions[guide_row].height = 32

    header_row = guide_row + 1
    for col, text in zip("A", ["Sheet Name"]):
        pass
    ws.merge_cells(f"A{header_row}:B{header_row}")
    style_col_header(ws[f"A{header_row}"], "Sheet Name")
    ws.merge_cells(f"C{header_row}:E{header_row}")
    style_col_header(ws[f"C{header_row}"], "Purpose")
    ws.merge_cells(f"F{header_row}:G{header_row}")
    style_col_header(ws[f"F{header_row}"], "Best Used For")
    style_col_header(ws[f"H{header_row}"], "How Often")

    sheet_guide = [
        ("\U0001F4CA Dashboard", "Command center with KPIs and progress", "Quick status checks", "Weekly"),
        ("\U0001F4C5 Monthly Budget", "Plan and track monthly spending limits", "Setting category budgets", "Monthly"),
        ("\U0001F4C6 Weekly Budget", "Break the month into 5 weekly views", "Tight week-to-week control", "Weekly"),
        ("\U0001F4B5 Paycheck Budget", "Allocate each paycheck to categories", "Paycheck-to-paycheck planning", "Per paycheck"),
        ("\U0001F4CB Biweekly Budget", "Plan two pay periods per month", "Biweekly pay schedules", "Biweekly"),
        ("\U0001F4B3 Expense Tracker", "Log every transaction", "Daily expense logging", "Daily"),
        ("\U0001F4B0 Income Tracker", "Log all income sources", "Tracking gross/net income", "Per income"),
        ("\U0001F3AF Savings Goals", "Track progress toward savings targets", "Big purchases & emergency funds", "Monthly"),
        ("\U0001F3E6 Debt Tracker", "Plan and track debt payoff", "Paying down loans & cards", "Monthly"),
        ("\U0001FAA3 Sinking Funds", "Save monthly for future expenses", "Irregular planned costs", "Monthly"),
        ("\U0001F4F1 Subscriptions", "Track recurring subscriptions", "Finding what to cancel", "Monthly"),
        ("\U0001F4C5 Bills Calendar", "Track due dates for every bill", "Never missing a payment", "Weekly"),
        ("\U0001F4C8 Annual Overview", "12-month financial summary", "Year-end review & trends", "Monthly"),
    ]
    for i, (name, purpose, best_for, freq) in enumerate(sheet_guide):
        row = header_row + 1 + i
        ws.merge_cells(f"A{row}:B{row}")
        ws[f"A{row}"] = name
        ws[f"A{row}"].font = Font(name=FONT_NAME, bold=True)
        ws.merge_cells(f"C{row}:E{row}")
        ws[f"C{row}"] = purpose
        ws.merge_cells(f"F{row}:G{row}")
        ws[f"F{row}"] = best_for
        ws[f"H{row}"] = freq
        ws[f"H{row}"].alignment = Alignment(horizontal="center")
        if i % 2 == 1:
            for col in "ABCDEFGH":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)
    apply_table_borders(ws, f"A{header_row}:H{header_row + len(sheet_guide)}")

    # Important tips
    tips_row = header_row + len(sheet_guide) + 2
    ws.merge_cells(f"A{tips_row}:H{tips_row}")
    style_section_header(ws[f"A{tips_row}"], "⚠️ IMPORTANT TIPS")
    ws.row_dimensions[tips_row].height = 32

    tips = [
        "Only edit cells with a YELLOW background — these are your inputs.",
        "GREEN and WHITE cells contain formulas — do not delete or overwrite them.",
        "Start a new copy of this file each month to keep history clean.",
        "Back up your file to Google Drive weekly.",
    ]
    for i, tip in enumerate(tips):
        row = tips_row + 1 + i
        ws.merge_cells(f"A{row}:H{row}")
        ws[f"A{row}"] = f"• {tip}"
        ws[f"A{row}"].font = Font(name=FONT_NAME, color=DARK_GRAY)

    delivery_row = tips_row + len(tips) + 2
    ws.merge_cells(f"A{delivery_row}:H{delivery_row + 4}")
    ws[f"A{delivery_row}"] = (
        "\U0001F4CC GOOGLE SHEETS USERS:\n"
        "Go to File → Import → Upload this file → Replace spreadsheet.\n"
        "Then visit each sheet and add charts using Insert → Chart for a richer view.\n\n"
        "\U0001F4CC EXCEL USERS:\n"
        "Open directly — all formatting and formulas work natively."
    )
    ws[f"A{delivery_row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws[f"A{delivery_row}"].font = Font(name=FONT_NAME, color=DARK_GRAY, italic=True)
    ws[f"A{delivery_row}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    ws.freeze_panes = "A3"
    finalize_sheet(ws)
    return ws


def build_monthly_budget(wb):
    print("✅ Building Monthly Budget...")
    ws = wb.create_sheet("Monthly Budget")
    ws.sheet_properties.tabColor = PRIMARY_GREEN

    set_col_widths(ws, {"A": 12, "B": 18, "C": 18, "D": 25, "E": 14, "F": 14,
                         "G": 14, "H": 10, "I": 16, "J": 25, "K": 14})

    ws.merge_cells("A1:K1")
    style_title(ws["A1"], "📅 MONTHLY BUDGET PLANNER")
    ws.row_dimensions[1].height = 40

    # --- Summary row ---
    ws.row_dimensions[2].height = 28
    summary = [
        ("A2", "Total Budgeted", "B2", "=SUM(E4:E103)"),
        ("C2", "Total Actual", "D2", "=SUM(F4:F103)"),
        ("E2", "Total Saved", "F2", "=B2-D2"),
    ]
    for label_cell, label, val_cell, formula in summary:
        ws[label_cell] = label
        style_kpi_label(ws[label_cell])
        ws[label_cell].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        ws[val_cell] = formula
        style_kpi(ws[val_cell], color=GOLD, currency=True)
        ws[val_cell].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws.merge_cells("G2:K2")
    ws["G2"] = ('=IF(F2>=0,"✅ Under Budget by "&TEXT(F2,"#,##0.00"),'
                 '"⚠️ Over Budget by "&TEXT(-F2,"#,##0.00"))')
    ws["G2"].font = Font(name=FONT_NAME, bold=True, size=12, color=DARK_TEXT)
    ws["G2"].alignment = Alignment(horizontal="left", vertical="center")
    ws["G2"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # --- Column headers ---
    headers = ["Date", "Category", "Subcategory", "Description", "Budgeted",
               "Actual", "Difference", "% Used", "Payment Method", "Notes", "Running Total"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        style_col_header(ws[f"{col}3"], h)
    ws.row_dimensions[3].height = 28

    # --- Example data (rows 4-15) ---
    today = date.today()
    examples = [
        ("Housing", "Rent", "Monthly rent payment", 1500, 1500, "Bank Transfer", ""),
        ("Food & Dining", "Groceries", "Weekly grocery run", 500, 540, "Debit Card", ""),
        ("Transportation", "Gas", "Fuel for car", 150, 165, "Credit Card", ""),
        ("Entertainment", "Streaming", "Netflix + Spotify", 30, 30, "Credit Card", ""),
        ("Shopping", "Clothing", "New work clothes", 100, 85, "Credit Card", ""),
        ("Healthcare", "Doctor", "Annual checkup copay", 50, 50, "Debit Card", ""),
        ("Utilities", "Electricity", "Power bill", 120, 135, "Bank Transfer", ""),
        ("Education", "Courses", "Online course", 50, 0, "Credit Card", "Not started yet"),
        ("Personal Care", "Haircuts", "Haircut", 40, 45, "Cash", ""),
        ("Savings", "General Savings", "Emergency fund transfer", 500, 500, "Bank Transfer", ""),
        ("Debt Payment", "Credit Card", "Minimum payment", 200, 200, "Bank Transfer", ""),
        ("Other", "Other", "Miscellaneous", 50, 60, "Cash", "Birthday gift"),
    ]
    for i, (cat, sub, desc, budgeted, actual, pay, note) in enumerate(examples):
        row = 4 + i
        ws[f"A{row}"] = date(today.year, today.month, min(28, i * 2 + 1))
        ws[f"B{row}"] = cat
        ws[f"C{row}"] = sub
        ws[f"D{row}"] = desc
        ws[f"E{row}"] = budgeted
        ws[f"F{row}"] = actual
        ws[f"I{row}"] = pay
        ws[f"J{row}"] = note

    # --- All 100 data rows: formulas, formats, borders ---
    for row in range(4, 104):
        ws[f"A{row}"].number_format = DATE_FMT
        ws[f"E{row}"].number_format = CURRENCY_FMT
        ws[f"F{row}"].number_format = CURRENCY_FMT
        ws[f"G{row}"] = f"=E{row}-F{row}"
        ws[f"G{row}"].number_format = CURRENCY_FMT
        ws[f"H{row}"] = f"=IFERROR(F{row}/E{row},0)"
        ws[f"H{row}"].number_format = PERCENT_FMT
        ws[f"K{row}"] = f"=SUM($F$4:F{row})"
        ws[f"K{row}"].number_format = CURRENCY_FMT
        for col in "ABCDEFGHIJK":
            apply_border(ws[f"{col}{row}"])
            if col in "ABCDEFIJ":
                ws[f"{col}{row}"].protection = Protection(locked=False)

    alternate_row_colors(ws, 4, 103, 1, 11)

    add_dropdown(ws, "B4:B103", "CATEGORIES")
    add_dropdown(ws, "C4:C103", "SUBCATEGORIES")
    add_dropdown(ws, "I4:I103", "PAYMENT_METHODS")

    cf_diff_green_red(ws, "G4:G103")
    cf_percent_used(ws, "H4:H103")
    ws.conditional_formatting.add(
        "A4:A103",
        CellIsRule(operator="equal", formula=["TODAY()"],
                   fill=PatternFill("solid", fgColor=YELLOW_INPUT)),
    )

    wb.defined_names.add(DefinedName("BudgetData", attr_text="'Monthly Budget'!$A$4:$K$103"))

    ws["A1"].comment = Comment(
        "Tip: this sheet covers one month. Copy it before starting a new month.",
        "Budget Planner")
    add_back_link(ws)
    ws.freeze_panes = "B4"
    finalize_sheet(ws)
    return ws


def build_weekly_budget(wb):
    print("✅ Building Weekly Budget...")
    ws = wb.create_sheet("Weekly Budget")
    ws.sheet_properties.tabColor = "B2A0A0"

    set_col_widths(ws, {"A": 22, "B": 14, "C": 14, "D": 14, "E": 14, "F": 14, "G": 16})

    ws.merge_cells("A1:G1")
    style_title(ws["A1"], "📆 WEEKLY BUDGET")
    ws.row_dimensions[1].height = 40

    style_col_header(ws["A2"], "Category")
    for i in range(5):
        col = get_column_letter(2 + i)  # B..F
        start_day = i * 7 + 1
        end_day = min(start_day + 6, 28)
        ws[f"{col}2"] = (
            f'="Week {i + 1}: "&TEXT(DATE(YEAR(TODAY()),MONTH(TODAY()),{start_day}),"MMM D")'
            f'&"–"&TEXT(DATE(YEAR(TODAY()),MONTH(TODAY()),{end_day}),"MMM D")'
        )
        style_col_header(ws[f"{col}2"])
    style_col_header(ws["G2"], "Monthly Total")
    ws.row_dimensions[2].height = 32

    # --- Income row ---
    ws["A3"] = "Income This Week"
    ws["A3"].font = Font(name=FONT_NAME, bold=True)
    apply_border(ws["A3"])
    for col in "BCDEF":
        style_input(ws[f"{col}3"], 1100)
        style_currency(ws[f"{col}3"])
    ws["G3"] = "=SUM(B3:F3)"
    style_currency(ws["G3"])
    apply_border(ws["G3"])
    ws["G3"].font = Font(name=FONT_NAME, bold=True)

    # --- Category rows ---
    categories = ["Housing", "Food & Dining", "Transportation", "Entertainment",
                   "Shopping", "Utilities", "Healthcare", "Personal", "Savings", "Other"]
    expense_start_row = 5
    row = expense_start_row
    for cat in categories:
        ws[f"A{row}"] = cat
        apply_border(ws[f"A{row}"])
        for col in "BCDEF":
            style_input(ws[f"{col}{row}"], 0)
            style_currency(ws[f"{col}{row}"])
        ws[f"G{row}"] = f"=SUM(B{row}:F{row})"
        style_currency(ws[f"G{row}"])
        apply_border(ws[f"G{row}"])
        row += 1
    expense_end_row = row - 1

    # --- Totals / Net / Remaining / Progress ---
    total_row = expense_end_row + 2
    ws[f"A{total_row}"] = "Total Expenses"
    ws[f"A{total_row}"].font = Font(name=FONT_NAME, bold=True)
    for col in "BCDEFG":
        ws[f"{col}{total_row}"] = f"=SUM({col}{expense_start_row}:{col}{expense_end_row})"
        style_currency(ws[f"{col}{total_row}"])
        ws[f"{col}{total_row}"].font = Font(name=FONT_NAME, bold=True)
        ws[f"{col}{total_row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    net_row = total_row + 1
    ws[f"A{net_row}"] = "Net (Income - Expenses)"
    for col in "BCDEFG":
        income_ref = "G3" if col == "G" else f"{col}3"
        ws[f"{col}{net_row}"] = f"={income_ref}-{col}{total_row}"
        style_currency(ws[f"{col}{net_row}"])

    remaining_row = net_row + 1
    ws[f"A{remaining_row}"] = "Remaining to Spend"
    for col in "BCDEFG":
        ws[f"{col}{remaining_row}"] = f"=MAX({col}{net_row},0)"
        style_currency(ws[f"{col}{remaining_row}"])

    progress_row = remaining_row + 1
    ws[f"A{progress_row}"] = "Budget Used"
    for col in "BCDEFG":
        income_ref = "G3" if col == "G" else f"{col}3"
        pct_expr = f"IFERROR({col}{total_row}/{income_ref},0)"
        ws[f"{col}{progress_row}"] = progress_bar_expr(pct_expr)
        ws[f"{col}{progress_row}"].font = Font(name=FONT_NAME, size=9)

    cf_diff_green_red(ws, f"B{net_row}:G{net_row}")

    # Alternating column shading for weeks 2 and 4
    for col in ["C", "E"]:
        for r in range(2, progress_row + 1):
            cell = ws[f"{col}{r}"]
            if cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = PatternFill("solid", fgColor=PALE_GREEN)

    apply_table_borders(ws, f"A2:G{progress_row}")

    add_back_link(ws)
    ws.freeze_panes = "B3"
    finalize_sheet(ws)
    return ws


def build_paycheck_budget(wb):
    print("✅ Building Paycheck Budget...")
    ws = wb.create_sheet("Paycheck Budget")
    ws.sheet_properties.tabColor = "C9B8B8"

    set_col_widths(ws, {"A": 12, "B": 13, "C": 9, "D": 13, "E": 11, "F": 12,
                         "G": 11, "H": 11, "I": 12, "J": 13, "K": 13, "L": 13,
                         "M": 12, "N": 16})

    ws.merge_cells("A1:N1")
    style_title(ws["A1"], "💵 PAYCHECK BUDGET")
    ws.row_dimensions[1].height = 40

    headers = ["Pay Date", "Gross Income", "Taxes %", "Net Income", "Bills",
               "Groceries", "Transport", "Savings", "Investments", "Debt Payment",
               "Entertainment", "Total Allocated", "Remaining", "Fully Allocated?"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        style_col_header(ws[f"{col}2"], h)
    ws.row_dimensions[2].height = 28

    today = date.today()
    data_start, data_end = 3, 14  # 12 paychecks
    from datetime import timedelta
    for i in range(12):
        row = data_start + i
        ws[f"A{row}"] = today + timedelta(days=7 * i)
        ws[f"B{row}"] = 1375
        ws[f"C{row}"] = 0.22
        ws[f"D{row}"] = f"=B{row}*(1-C{row})"
        ws[f"E{row}"] = 500
        ws[f"F{row}"] = 150
        ws[f"G{row}"] = 80
        ws[f"H{row}"] = 150
        ws[f"I{row}"] = 50
        ws[f"J{row}"] = 50
        ws[f"K{row}"] = 40
        ws[f"L{row}"] = f"=SUM(E{row}:K{row})"
        ws[f"M{row}"] = f"=D{row}-L{row}"
        ws[f"N{row}"] = f'=IF(ROUND(M{row},2)=0,"✅",IF(M{row}>0,"✅ "&TEXT(M{row},"#,##0.00")&" left","⚠️ "&TEXT(-M{row},"#,##0.00")&" over"))'

        ws[f"A{row}"].number_format = DATE_FMT
        for col in "BDEFGHIJKLM":
            ws[f"{col}{row}"].number_format = CURRENCY_FMT
        ws[f"C{row}"].number_format = PERCENT_FMT
        for col in "ABCEFGHIJK":
            ws[f"{col}{row}"].protection = Protection(locked=False)
        for col in "ABCDEFGHIJKLMN":
            apply_border(ws[f"{col}{row}"])

    alternate_row_colors(ws, data_start, data_end, 1, 14)
    cf_diff_green_red(ws, f"M{data_start}:M{data_end}")

    # --- Summary rows ---
    totals_row = data_end + 2
    avg_row = totals_row + 1
    rate_row = avg_row + 1

    ws[f"A{totals_row}"] = "Totals"
    ws[f"A{avg_row}"] = "Average per Paycheck"
    ws[f"A{rate_row}"] = "Savings Rate"
    for r in (totals_row, avg_row, rate_row):
        ws[f"A{r}"].font = Font(name=FONT_NAME, bold=True)
        ws[f"A{r}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    for col in "BDEFGHIJKLM":
        ws[f"{col}{totals_row}"] = f"=SUM({col}{data_start}:{col}{data_end})"
        ws[f"{col}{totals_row}"].number_format = CURRENCY_FMT
        ws[f"{col}{totals_row}"].font = Font(name=FONT_NAME, bold=True)
        ws[f"{col}{totals_row}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

        ws[f"{col}{avg_row}"] = f"=AVERAGE({col}{data_start}:{col}{data_end})"
        ws[f"{col}{avg_row}"].number_format = CURRENCY_FMT
        ws[f"{col}{avg_row}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    ws.merge_cells(f"B{rate_row}:N{rate_row}")
    ws[f"B{rate_row}"] = f"=(H{totals_row}+I{totals_row})/D{totals_row}"
    style_percent(ws[f"B{rate_row}"])
    ws[f"B{rate_row}"].font = Font(name=FONT_NAME, bold=True, color=POSITIVE_TEXT, size=12)
    ws[f"B{rate_row}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # --- 50/30/20 allocation breakdown ---
    alloc_header_row = rate_row + 2
    ws.merge_cells(f"A{alloc_header_row}:N{alloc_header_row}")
    style_section_header(ws[f"A{alloc_header_row}"], "📐 50/30/20 ALLOCATION CHECK")
    ws.row_dimensions[alloc_header_row].height = 32

    col_header_row = alloc_header_row + 1
    ws.merge_cells(f"A{col_header_row}:D{col_header_row}")
    style_col_header(ws[f"A{col_header_row}"], "Allocation Group")
    style_col_header(ws[f"E{col_header_row}"], "Your %")
    style_col_header(ws[f"F{col_header_row}"], "Target %")
    ws.merge_cells(f"G{col_header_row}:N{col_header_row}")
    style_col_header(ws[f"G{col_header_row}"], "Status")

    groups = [
        ("Needs (Bills, Groceries, Transport)", f"(E{totals_row}+F{totals_row}+G{totals_row})/D{totals_row}", 0.50),
        ("Wants (Entertainment)", f"K{totals_row}/D{totals_row}", 0.30),
        ("Savings & Debt (Savings, Investments, Debt)", f"(H{totals_row}+I{totals_row}+J{totals_row})/D{totals_row}", 0.20),
    ]
    for i, (label, expr, target) in enumerate(groups):
        row = col_header_row + 1 + i
        ws.merge_cells(f"A{row}:D{row}")
        ws[f"A{row}"] = label
        ws[f"E{row}"] = f"={expr}"
        style_percent(ws[f"E{row}"])
        ws[f"F{row}"] = target
        style_percent(ws[f"F{row}"])
        ws.merge_cells(f"G{row}:N{row}")
        ws[f"G{row}"] = f'=IF(E{row}<=F{row}*1.1,"✅ On Track","⚠️ Above Target")'
        for col in "ABCDEFGHIJKLMN":
            apply_border(ws[f"{col}{row}"])
        if i % 2 == 1:
            for col in "ABCDEFGHIJKLMN":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    add_back_link(ws)
    ws.freeze_panes = "B3"
    finalize_sheet(ws)
    return ws


def build_biweekly_budget(wb):
    print("✅ Building Biweekly Budget...")
    ws = wb.create_sheet("Biweekly Budget")
    ws.sheet_properties.tabColor = "D4EDEA"

    set_col_widths(ws, {"A": 36, "B": 18, "C": 18, "D": 18})

    ws.merge_cells("A1:D1")
    style_title(ws["A1"], "📋 BIWEEKLY BUDGET")
    ws.row_dimensions[1].height = 40

    style_col_header(ws["A2"], "Category")
    style_col_header(ws["B2"], "PAYCHECK 1 (1st–15th)")
    style_col_header(ws["C2"], "PAYCHECK 2 (16th–End)")
    style_col_header(ws["D2"], "Monthly Total")
    ws.row_dimensions[2].height = 32

    income_row, rollover_row, available_row = 3, 4, 5
    section_row = 6
    fixed_row, variable_row, savings_row, invest_row, debt_row = 7, 8, 9, 10, 11
    total_exp_row, remaining_row = 12, 13

    rows = {
        income_row: ("Net Income (Paycheck)", 1900, 1900),
        rollover_row: ("Rollover from Previous Period", 0, None),
        available_row: ("Available Funds", None, None),
        fixed_row: ("Fixed Bills (Rent, Insurance, Subscriptions)", 1500, 200),
        variable_row: ("Variable Expenses (Food, Gas, Entertainment)", 350, 350),
        savings_row: ("Savings", 150, 150),
        invest_row: ("Investments", 50, 50),
        debt_row: ("Debt Payments", 100, 100),
        total_exp_row: ("Total Expenses", None, None),
        remaining_row: ("Remaining Balance", None, None),
    }

    for row, (label, b_val, c_val) in rows.items():
        ws[f"A{row}"] = label
        if row in (available_row, total_exp_row, remaining_row):
            ws[f"A{row}"].font = Font(name=FONT_NAME, bold=True)
        apply_border(ws[f"A{row}"])

    # Section header for expenses
    ws.merge_cells(f"A{section_row}:D{section_row}")
    style_section_header(ws[f"A{section_row}"], "💸 EXPENSES")
    ws[f"A{section_row}"] = "💸 EXPENSES"

    # Income (inputs)
    style_input(ws[f"B{income_row}"], 1900)
    style_input(ws[f"C{income_row}"], 1900)
    style_currency(ws[f"B{income_row}"])
    style_currency(ws[f"C{income_row}"])
    ws[f"D{income_row}"] = f"=B{income_row}+C{income_row}"
    style_currency(ws[f"D{income_row}"])
    apply_border(ws[f"D{income_row}"])

    # Rollover
    ws[f"B{rollover_row}"] = 0
    style_currency(ws[f"B{rollover_row}"])
    apply_border(ws[f"B{rollover_row}"])
    ws[f"C{rollover_row}"] = f"=B{remaining_row}"
    style_currency(ws[f"C{rollover_row}"])
    apply_border(ws[f"C{rollover_row}"])
    ws[f"D{rollover_row}"] = "—"
    ws[f"D{rollover_row}"].alignment = Alignment(horizontal="center")
    apply_border(ws[f"D{rollover_row}"])

    # Available funds
    for col in "BCD":
        ws[f"{col}{available_row}"] = f"={col}{income_row}+{col}{rollover_row}" if col != "D" else f"=B{available_row}+C{available_row}"
        style_currency(ws[f"{col}{available_row}"])
        ws[f"{col}{available_row}"].font = Font(name=FONT_NAME, bold=True)
        ws[f"{col}{available_row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)
        apply_border(ws[f"{col}{available_row}"])

    # Expense input rows
    for row in (fixed_row, variable_row, savings_row, invest_row, debt_row):
        b_val = rows[row][1]
        c_val = rows[row][2]
        style_input(ws[f"B{row}"], b_val)
        style_input(ws[f"C{row}"], c_val)
        style_currency(ws[f"B{row}"])
        style_currency(ws[f"C{row}"])
        ws[f"D{row}"] = f"=B{row}+C{row}"
        style_currency(ws[f"D{row}"])
        apply_border(ws[f"D{row}"])

    # Total expenses
    for col in "BCD":
        ws[f"{col}{total_exp_row}"] = f"=SUM({col}{fixed_row}:{col}{debt_row})"
        style_currency(ws[f"{col}{total_exp_row}"])
        ws[f"{col}{total_exp_row}"].font = Font(name=FONT_NAME, bold=True)
        apply_border(ws[f"{col}{total_exp_row}"])

    # Remaining balance
    ws[f"B{remaining_row}"] = f"=B{available_row}-B{total_exp_row}"
    ws[f"C{remaining_row}"] = f"=C{available_row}-C{total_exp_row}"
    ws[f"D{remaining_row}"] = f"=C{remaining_row}"
    for col in "BCD":
        style_currency(ws[f"{col}{remaining_row}"])
        ws[f"{col}{remaining_row}"].font = Font(name=FONT_NAME, bold=True)
        apply_border(ws[f"{col}{remaining_row}"])

    cf_diff_green_red(ws, f"B{remaining_row}:D{remaining_row}")
    apply_table_borders(ws, f"A2:D{remaining_row}")

    note_row = remaining_row + 2
    ws.merge_cells(f"A{note_row}:D{note_row + 1}")
    ws[f"A{note_row}"] = (
        "💡 Any leftover (or shortfall) from Paycheck 1 automatically rolls into "
        "Paycheck 2's available funds, so you always see a true running balance."
    )
    ws[f"A{note_row}"].alignment = Alignment(wrap_text=True, vertical="top")
    ws[f"A{note_row}"].font = Font(name=FONT_NAME, italic=True, color=DARK_GRAY)
    ws[f"A{note_row}"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    add_back_link(ws)
    ws.freeze_panes = "B3"
    finalize_sheet(ws)
    return ws


def build_expense_tracker(wb):
    print("✅ Building Expense Tracker...")
    ws = wb.create_sheet("Expense Tracker")
    ws.sheet_properties.tabColor = SOFT_RED

    set_col_widths(ws, {"A": 12, "B": 22, "C": 18, "D": 18, "E": 14, "F": 16,
                         "G": 14, "H": 12, "I": 25, "J": 14, "K": 8, "L": 8, "M": 8})

    ws.merge_cells("A1:M1")
    style_title(ws["A1"], "💳 EXPENSE TRACKER")
    ws.row_dimensions[1].height = 40

    data_first, data_last = 4, 503

    # --- Summary stats bar ---
    ws.row_dimensions[2].height = 28
    stats = [
        ("A2", "Total Spent", "B2", f"=SUM(E{data_first}:E{data_last})", True),
        ("C2", "Needs Total", "D2", f'=SUMIF(G{data_first}:G{data_last},"Need",E{data_first}:E{data_last})', True),
        ("E2", "Wants Total", "F2", f'=SUMIF(G{data_first}:G{data_last},"Want",E{data_first}:E{data_last})', True),
        ("G2", "Largest Purchase", "H2", f"=MAX(E{data_first}:E{data_last})", True),
        ("I2", "Transactions", "J2", f"=COUNTA(B{data_first}:B{data_last})", False),
    ]
    for label_cell, label, val_cell, formula, currency in stats:
        ws[label_cell] = label
        style_kpi_label(ws[label_cell])
        ws[label_cell].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        ws[val_cell] = formula
        style_kpi(ws[val_cell], color=GOLD if currency else DARK_TEXT, currency=currency)
        ws[val_cell].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws.merge_cells("K2:M2")
    ws["K2"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # --- Column headers ---
    headers = ["Date", "Merchant", "Category", "Subcategory", "Amount", "Payment Method",
               "Need or Want", "Recurring?", "Notes", "Month", "Week#", "Month#", "Year"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        style_col_header(ws[f"{col}3"], h)
    ws.row_dimensions[3].height = 28

    # --- Example transactions (rows 4-20) ---
    today = date.today()
    examples = [
        (1, "Whole Foods", "Food & Dining", "Groceries", 125.40, "Debit Card", "Need", "No", ""),
        (2, "Amazon", "Shopping", "Electronics", 89.99, "Credit Card", "Want", "No", ""),
        (2, "Shell Gas", "Transportation", "Gas", 52.30, "Credit Card", "Need", "No", ""),
        (3, "Netflix", "Entertainment", "Streaming", 15.49, "Credit Card", "Want", "Yes", ""),
        (3, "Spotify", "Entertainment", "Streaming", 11.99, "Credit Card", "Want", "Yes", ""),
        (4, "Target", "Shopping", "Clothing", 64.20, "Debit Card", "Want", "No", ""),
        (5, "CVS Pharmacy", "Healthcare", "Pharmacy", 23.75, "Debit Card", "Need", "No", ""),
        (6, "Uber", "Transportation", "Public Transit", 18.50, "Credit Card", "Want", "No", ""),
        (7, "Chipotle", "Food & Dining", "Restaurants", 13.85, "Cash", "Want", "No", ""),
        (7, "Starbucks", "Food & Dining", "Coffee", 6.45, "Debit Card", "Want", "No", "Daily habit"),
        (8, "Planet Fitness", "Personal Care", "Gym", 29.99, "Bank Transfer", "Need", "Yes", ""),
        (9, "Apple", "Shopping", "Electronics", 249.00, "Credit Card", "Want", "No", "New earbuds"),
        (10, "Whole Foods", "Food & Dining", "Groceries", 142.10, "Debit Card", "Need", "No", ""),
        (11, "Shell Gas", "Transportation", "Gas", 48.75, "Credit Card", "Need", "No", ""),
        (12, "Electric Co.", "Utilities", "Electricity", 135.00, "Bank Transfer", "Need", "Yes", ""),
        (12, "Verizon", "Utilities", "Phone", 85.00, "Bank Transfer", "Need", "Yes", ""),
        (13, "Amazon", "Shopping", "Other", 34.99, "Credit Card", "Want", "No", "Phone case"),
    ]
    for i, (day, merchant, cat, sub, amount, pay, nw, recur, note) in enumerate(examples):
        row = data_first + i
        ws[f"A{row}"] = date(today.year, today.month, min(day, 28))
        ws[f"B{row}"] = merchant
        ws[f"C{row}"] = cat
        ws[f"D{row}"] = sub
        ws[f"E{row}"] = amount
        ws[f"F{row}"] = pay
        ws[f"G{row}"] = nw
        ws[f"H{row}"] = recur
        ws[f"I{row}"] = note

    # --- Formats / helper formulas for all 500 rows ---
    for row in range(data_first, data_last + 1):
        ws[f"A{row}"].number_format = DATE_FMT
        ws[f"E{row}"].number_format = CURRENCY_FMT
        ws[f"E{row}"].font = Font(name=FONT_NAME, bold=True)
        ws[f"J{row}"] = f'=IF(A{row}="","",TEXT(A{row},"MMMM YYYY"))'
        ws[f"K{row}"] = f'=IF(A{row}="","",WEEKNUM(A{row}))'
        ws[f"L{row}"] = f'=IF(A{row}="","",MONTH(A{row}))'
        ws[f"M{row}"] = f'=IF(A{row}="","",YEAR(A{row}))'
        for col in "ABCDEFGHIJKLM":
            apply_border(ws[f"{col}{row}"])
            if col in "ABCDEFGHI":
                ws[f"{col}{row}"].protection = Protection(locked=False)

    for col in ("J", "K", "L", "M"):
        ws.column_dimensions[col].hidden = True

    alternate_row_colors(ws, data_first, data_last, 1, 13)

    add_dropdown(ws, f"C{data_first}:C{data_last}", "CATEGORIES")
    add_dropdown(ws, f"D{data_first}:D{data_last}", "SUBCATEGORIES")
    add_dropdown(ws, f"F{data_first}:F{data_last}", "PAYMENT_METHODS")
    add_dropdown(ws, f"G{data_first}:G{data_last}", "NEED_WANT")
    add_dropdown(ws, f"H{data_first}:H{data_last}", "YES_NO")

    # --- Conditional formatting ---
    nw_range = f"G{data_first}:G{data_last}"
    ws.conditional_formatting.add(nw_range, CellIsRule(operator="equal", formula=['"Need"'],
                                                         fill=PatternFill("solid", fgColor=LIGHT_BLUE)))
    ws.conditional_formatting.add(nw_range, CellIsRule(operator="equal", formula=['"Want"'],
                                                         fill=PatternFill("solid", fgColor=LIGHT_ORANGE)))
    ws.conditional_formatting.add(nw_range, CellIsRule(operator="equal", formula=['"Investment"'],
                                                         fill=PatternFill("solid", fgColor=LIGHT_GREEN)))

    ws.conditional_formatting.add(
        f"H{data_first}:H{data_last}",
        CellIsRule(operator="equal", formula=['"Yes"'], fill=PatternFill("solid", fgColor=LAVENDER)),
    )

    top10 = Rule(type="top10", rank=10,
                  dxf=DifferentialStyle(fill=PatternFill("solid", fgColor=GOLD)))
    ws.conditional_formatting.add(f"E{data_first}:E{data_last}", top10)

    bold_row_rule = FormulaRule(formula=[f"$E{data_first}>200"],
                                 font=Font(name=FONT_NAME, bold=True))
    ws.conditional_formatting.add(f"A{data_first}:I{data_last}", bold_row_rule)

    ws["A1"].comment = Comment(
        "Tip: Use Data > Sort by Date (column A) to keep transactions in order.",
        "Budget Planner")

    # --- Category Summary table ---
    summary_row = data_last + 3
    ws.merge_cells(f"A{summary_row}:E{summary_row}")
    style_section_header(ws[f"A{summary_row}"], "📊 CATEGORY SUMMARY")
    cs_header_row = summary_row + 1
    for col, h in zip("ABCDE", ["Category", "Total Spent", "# Transactions", "Avg Transaction", "% of Total"]):
        style_col_header(ws[f"{col}{cs_header_row}"], h)
    for i, cat in enumerate(CATEGORIES_LIST):
        row = cs_header_row + 1 + i
        ws[f"A{row}"] = cat
        ws[f"B{row}"] = f"=SUMIF(C{data_first}:C{data_last},A{row},E{data_first}:E{data_last})"
        ws[f"C{row}"] = f"=COUNTIF(C{data_first}:C{data_last},A{row})"
        ws[f"D{row}"] = f"=IFERROR(B{row}/C{row},0)"
        ws[f"E{row}"] = f"=IFERROR(B{row}/SUM(E{data_first}:E{data_last}),0)"
        style_currency(ws[f"B{row}"])
        style_currency(ws[f"D{row}"])
        style_percent(ws[f"E{row}"])
        for col in "ABCDE":
            apply_border(ws[f"{col}{row}"])
        if i % 2 == 1:
            for col in "ABCDE":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    # --- Top 5 Largest Expenses table ---
    ws.merge_cells(f"G{summary_row}:J{summary_row}")
    style_section_header(ws[f"G{summary_row}"], "🏆 TOP 5 EXPENSES")
    for col, h in zip("GHIJ", ["Rank", "Merchant", "Category", "Amount"]):
        style_col_header(ws[f"{col}{cs_header_row}"], h)
    for rank in range(1, 6):
        row = cs_header_row + rank
        ws[f"G{row}"] = rank
        ws[f"G{row}"].alignment = Alignment(horizontal="center")
        ws[f"J{row}"] = f"=LARGE($E${data_first}:$E${data_last},{rank})"
        ws[f"H{row}"] = (f'=IFERROR(INDEX($B${data_first}:$B${data_last},'
                         f'MATCH(J{row},$E${data_first}:$E${data_last},0)),"")')
        ws[f"I{row}"] = (f'=IFERROR(INDEX($C${data_first}:$C${data_last},'
                         f'MATCH(J{row},$E${data_first}:$E${data_last},0)),"")')
        style_currency(ws[f"J{row}"])
        for col in "GHIJ":
            apply_border(ws[f"{col}{row}"])
        if rank % 2 == 0:
            for col in "GHIJ":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    wb.defined_names.add(DefinedName(
        "ExpenseData", attr_text=f"'Expense Tracker'!$A${data_first}:$M${data_last}"))

    add_back_link(ws)
    ws.freeze_panes = "B4"
    finalize_sheet(ws)
    return ws


def build_income_tracker(wb):
    print("✅ Building Income Tracker...")
    ws = wb.create_sheet("Income Tracker")
    ws.sheet_properties.tabColor = GOLD

    set_col_widths(ws, {"A": 12, "B": 20, "C": 28, "D": 14, "E": 11, "F": 13,
                         "G": 14, "H": 16, "I": 14, "J": 9, "K": 8})

    ws.merge_cells("A1:K1")
    style_title(ws["A1"], "💰 INCOME TRACKER")
    ws.row_dimensions[1].height = 40

    data_first, data_last = 4, 103

    # --- Summary stats bar ---
    ws.row_dimensions[2].height = 28
    stats = [
        ("A2", "Total Gross", "B2", f"=SUM(D{data_first}:D{data_last})", True),
        ("C2", "Tax Reserved", "D2", f"=SUM(F{data_first}:F{data_last})", True),
        ("E2", "Total Net", "F2", f"=SUM(G{data_first}:G{data_last})", True),
        ("G2", "Avg per Entry", "H2", f"=IFERROR(AVERAGE(D{data_first}:D{data_last}),0)", True),
    ]
    for label_cell, label, val_cell, formula, currency in stats:
        ws[label_cell] = label
        style_kpi_label(ws[label_cell])
        ws[label_cell].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
        ws[val_cell] = formula
        style_kpi(ws[val_cell], color=GOLD, currency=currency)
        ws[val_cell].fill = PatternFill("solid", fgColor=LIGHT_GRAY)
    ws.merge_cells("I2:K2")
    ws["I2"].fill = PatternFill("solid", fgColor=LIGHT_GRAY)

    # --- Column headers ---
    headers = ["Date", "Source", "Description", "Amount", "Tax Rate %", "Tax Reserve",
               "Net After Tax", "Month", "Category Type", "Month#", "Year"]
    for i, h in enumerate(headers):
        col = get_column_letter(i + 1)
        style_col_header(ws[f"{col}3"], h)
    ws.row_dimensions[3].height = 28

    # --- Example data: 6 months of biweekly salary + 2 extra sources ---
    def shift_month(d, months_back):
        total = d.year * 12 + (d.month - 1) - months_back
        return total // 12, total % 12 + 1

    today = date.today()
    examples = []
    for months_ago in range(5, -1, -1):
        for day in (1, 15):
            y, m = shift_month(today, months_ago)
            examples.append((date(y, m, day), "Salary", "Biweekly paycheck", 2750, 0.22, "Active"))
    y, m = shift_month(today, 4)
    examples.append((date(y, m, 20), "Freelance/Contract", "Web design project", 600, 0.25, "Active"))
    y, m = shift_month(today, 2)
    examples.append((date(y, m, 25), "Dividends", "Investment dividend payout", 45, 0.15, "Passive"))

    for i, (d, source, desc, amount, tax_rate, cat_type) in enumerate(examples):
        row = data_first + i
        ws[f"A{row}"] = d
        ws[f"B{row}"] = source
        ws[f"C{row}"] = desc
        ws[f"D{row}"] = amount
        ws[f"E{row}"] = tax_rate
        ws[f"I{row}"] = cat_type

    # --- Formulas / formats for all 100 rows ---
    for row in range(data_first, data_last + 1):
        ws[f"A{row}"].number_format = DATE_FMT
        ws[f"D{row}"].number_format = CURRENCY_FMT
        ws[f"E{row}"].number_format = PERCENT_FMT
        ws[f"F{row}"] = f"=D{row}*E{row}"
        ws[f"F{row}"].number_format = CURRENCY_FMT
        ws[f"G{row}"] = f"=D{row}-F{row}"
        ws[f"G{row}"].number_format = CURRENCY_FMT
        ws[f"H{row}"] = f'=IF(A{row}="","",TEXT(A{row},"MMMM YYYY"))'
        ws[f"J{row}"] = f'=IF(A{row}="","",MONTH(A{row}))'
        ws[f"K{row}"] = f'=IF(A{row}="","",YEAR(A{row}))'
        for col in "ABCDEFGHIJK":
            apply_border(ws[f"{col}{row}"])
            if col in "ABCDEI":
                ws[f"{col}{row}"].protection = Protection(locked=False)

    for col in ("J", "K"):
        ws.column_dimensions[col].hidden = True

    alternate_row_colors(ws, data_first, data_last, 1, 11)

    add_dropdown(ws, f"B{data_first}:B{data_last}", "INCOME_SOURCES")
    add_dropdown(ws, f"I{data_first}:I{data_last}", "CATEGORY_TYPE")

    # --- Monthly Summary table ---
    summary_row = data_last + 3
    ws.merge_cells(f"A{summary_row}:E{summary_row}")
    style_section_header(ws[f"A{summary_row}"], "📅 MONTHLY SUMMARY")
    header_row = summary_row + 1
    for col, h in zip("ABCDE", ["Month", "Gross Income", "Tax Reserve", "Net Income", "# Sources"]):
        style_col_header(ws[f"{col}{header_row}"], h)

    month_rows_start = header_row + 1
    for m in range(1, 13):
        row = month_rows_start + m - 1
        ws[f"A{row}"] = date(2000, m, 1).strftime("%B")
        ws[f"B{row}"] = (f"=SUMIFS(D{data_first}:D{data_last},"
                         f"J{data_first}:J{data_last},{m},"
                         f"K{data_first}:K{data_last},YEAR(TODAY()))")
        ws[f"C{row}"] = (f"=SUMIFS(F{data_first}:F{data_last},"
                         f"J{data_first}:J{data_last},{m},"
                         f"K{data_first}:K{data_last},YEAR(TODAY()))")
        ws[f"D{row}"] = f"=B{row}-C{row}"
        ws[f"E{row}"] = (f"=COUNTIFS(J{data_first}:J{data_last},{m},"
                         f"K{data_first}:K{data_last},YEAR(TODAY()))")
        for col in "BCD":
            style_currency(ws[f"{col}{row}"])
        for col in "ABCDE":
            apply_border(ws[f"{col}{row}"])
        if m % 2 == 0:
            for col in "ABCDE":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)
    month_rows_end = month_rows_start + 11

    # --- Annual Summary + Income by Source ---
    bottom_section_row = month_rows_end + 2
    ws.merge_cells(f"A{bottom_section_row}:B{bottom_section_row}")
    style_section_header(ws[f"A{bottom_section_row}"], "📈 ANNUAL SUMMARY")
    ws.merge_cells(f"D{bottom_section_row}:G{bottom_section_row}")
    style_section_header(ws[f"D{bottom_section_row}"], "💼 INCOME BY SOURCE")

    annual_items = [
        ("Total Gross Income", f"=SUM(B{month_rows_start}:B{month_rows_end})", True),
        ("Total Tax Reserved", f"=SUM(C{month_rows_start}:C{month_rows_end})", True),
        ("Total Net Income", f"=SUM(D{month_rows_start}:D{month_rows_end})", True),
        ("Best Income Month", f"=INDEX(A{month_rows_start}:A{month_rows_end},"
                               f"MATCH(MAX(B{month_rows_start}:B{month_rows_end}),"
                               f"B{month_rows_start}:B{month_rows_end},0))", False),
        ("Average Monthly Income", f"=IFERROR(AVERAGEIF(B{month_rows_start}:B{month_rows_end},\">0\"),0)", True),
    ]
    for i, (label, formula, currency) in enumerate(annual_items):
        row = bottom_section_row + 1 + i
        ws[f"A{row}"] = label
        ws[f"A{row}"].font = Font(name=FONT_NAME, bold=True)
        ws[f"B{row}"] = formula
        if currency:
            style_currency(ws[f"B{row}"])
        for col in "AB":
            apply_border(ws[f"{col}{row}"])

    for i, source in enumerate(INCOME_SOURCES_LIST):
        row = bottom_section_row + 1 + i
        ws[f"D{row}"] = source
        ws[f"E{row}"] = f"=SUMIF(B{data_first}:B{data_last},D{row},D{data_first}:D{data_last})"
        ws[f"F{row}"] = f"=IFERROR(E{row}/SUM(D{data_first}:D{data_last}),0)"
        style_currency(ws[f"E{row}"])
        style_percent(ws[f"F{row}"])
        for col in "DEF":
            apply_border(ws[f"{col}{row}"])
        if i % 2 == 1:
            for col in "DEFG":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    wb.defined_names.add(DefinedName(
        "IncomeData", attr_text=f"'Income Tracker'!$A${data_first}:$K${data_last}"))

    add_back_link(ws)
    ws.freeze_panes = "B4"
    finalize_sheet(ws)
    return ws


def build_savings_goals(wb):
    print("✅ Building Savings Goals...")
    ws = wb.create_sheet("Savings Goals")
    ws.sheet_properties.tabColor = "C9A84C"
    set_col_widths(ws, {"A": 16, "B": 14, "C": 22, "D": 14, "E": 20, "F": 14, "G": 12, "H": 14})

    ws.merge_cells("A1:H1")
    style_title(ws["A1"], "\U0001F3AF Savings Goals")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = ("Set up to 6 savings goals below. Enter your saved amount, monthly "
                 "contribution, and target date in the yellow cells — progress bars, "
                 "remaining amounts, and projections update automatically.")
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=DARK_GRAY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    ws.merge_cells("A3:H3")
    style_section_header(ws["A3"], "\U0001F49A YOUR GOALS")

    today = date.today()
    christmas = date(today.year, 12, 25)
    if christmas < today:
        christmas = date(today.year + 1, 12, 25)

    goals = [
        ("Emergency Fund", 10000, 6500, 300, today + timedelta(days=365)),
        ("Vacation", 3000, 2400, 150, today + timedelta(days=120)),
        ("New Car", 8000, 2000, 200, today + timedelta(days=730)),
        ("Down Payment", 25000, 9000, 400, today + timedelta(days=1095)),
        ("Investment Fund", 5000, 4200, 250, today + timedelta(days=90)),
        ("Christmas Gifts", 1500, 900, 100, christmas),
    ]

    saved_cells = []
    target_cells = []

    for i, (name, target, saved, contribution, deadline) in enumerate(goals):
        r0 = 4 + i * 5
        r1, r2, r3, r4 = r0 + 1, r0 + 2, r0 + 3, r0 + 4

        # Row 0: Goal name + target amount
        ws.merge_cells(f"A{r0}:D{r0}")
        name_cell = ws[f"A{r0}"]
        name_cell.value = f"\U0001F3AF {name}"
        name_cell.font = Font(name=FONT_NAME, size=13, bold=True, color=DARK_TEXT)
        name_cell.fill = PatternFill("solid", fgColor=PALE_GREEN)
        name_cell.alignment = Alignment(horizontal="left", vertical="center")

        label_cell = ws[f"E{r0}"]
        label_cell.value = "Target Amount:"
        label_cell.font = Font(name=FONT_NAME, bold=True, color=DARK_GRAY)
        label_cell.alignment = Alignment(horizontal="right", vertical="center")
        label_cell.fill = PatternFill("solid", fgColor=PALE_GREEN)

        ws.merge_cells(f"F{r0}:H{r0}")
        target_cell = ws[f"F{r0}"]
        style_input(target_cell, target)
        style_currency(target_cell)
        target_cell.font = Font(name=FONT_NAME, size=12, bold=True)

        for col in "ABCDEFGH":
            apply_border(ws[f"{col}{r0}"])

        # Row 1: Saved | Monthly contribution | Target date | Remaining
        fields = [
            ("A", "\U0001F4B0 Saved:", "B", saved, "currency"),
            ("C", "\U0001F4C6 Monthly Contribution:", "D", contribution, "currency"),
            ("E", "\U0001F5D3️ Target Date:", "F", deadline, "date"),
            ("G", "Remaining:", "H", None, "formula"),
        ]
        for label_col, label_text, value_col, value, kind in fields:
            lc = ws[f"{label_col}{r1}"]
            lc.value = label_text
            lc.font = Font(name=FONT_NAME, size=9, bold=True, color=DARK_GRAY)
            lc.alignment = Alignment(horizontal="right", vertical="center")
            apply_border(lc)

            vc = ws[f"{value_col}{r1}"]
            if kind == "formula":
                vc.value = f"=F{r0}-B{r1}"
                style_currency(vc)
                apply_border(vc)
            else:
                style_input(vc, value)
                if kind == "currency":
                    style_currency(vc)
                else:
                    style_date_cell(vc)

        # Row 2: Progress bar
        ws.merge_cells(f"A{r2}:H{r2}")
        pct_expr = f"IFERROR(MIN(B{r1}/F{r0},1),0)"
        bar_cell = ws[f"A{r2}"]
        bar_cell.value = progress_bar_expr(pct_expr)
        bar_cell.font = Font(name=FONT_NAME, size=12, color=POSITIVE_TEXT)
        bar_cell.alignment = Alignment(horizontal="left", vertical="center")
        apply_border(bar_cell)

        # Row 3: Projection text
        ws.merge_cells(f"A{r3}:H{r3}")
        info_cell = ws[f"A{r3}"]
        info_cell.value = (
            f'=IF(B{r1}>=F{r0},"\U0001F389 Goal reached — congratulations!",'
            f'"\U0001F4C5 On track to reach this goal around "&'
            f'IF(D{r1}<=0,"N/A (set a monthly contribution)",'
            f'TEXT(TODAY()+(F{r0}-B{r1})/D{r1}*30,"MMMM YYYY"))&'
            f'".   \U0001F4A1 To hit your "&TEXT(F{r1},"MMM YYYY")&" target date, save "&'
            f'TEXT(IFERROR((F{r0}-B{r1})/MAX(DATEDIF(TODAY(),F{r1},"M"),1),F{r0}-B{r1}),"$#,##0")&"/month.")'
        )
        info_cell.font = Font(name=FONT_NAME, size=9, italic=True, color=DARK_GRAY)
        info_cell.alignment = Alignment(horizontal="left", vertical="center")
        apply_border(info_cell)

        # Spacer row
        ws.row_dimensions[r4].height = 8

        # Color-code the whole card by progress
        block_range = f"A{r0}:H{r3}"
        ws.conditional_formatting.add(
            block_range,
            FormulaRule(formula=[f"AND(F{r1}<TODAY(),B{r1}<F{r0})"],
                        fill=PatternFill("solid", fgColor=LIGHT_RED), stopIfTrue=True),
        )
        ws.conditional_formatting.add(
            block_range,
            FormulaRule(formula=[f"B{r1}/F{r0}>=0.75"],
                        fill=PatternFill("solid", fgColor=LIGHT_GREEN), stopIfTrue=True),
        )
        ws.conditional_formatting.add(
            block_range,
            FormulaRule(formula=[f"AND(B{r1}/F{r0}>=0.5,B{r1}/F{r0}<0.75)"],
                        fill=PatternFill("solid", fgColor=GOLD), stopIfTrue=True),
        )
        ws.conditional_formatting.add(
            block_range,
            FormulaRule(formula=[f"B{r1}/F{r0}<0.5"],
                        fill=PatternFill("solid", fgColor=LIGHT_BLUE), stopIfTrue=True),
        )

        saved_cells.append(f"B{r1}")
        target_cells.append(f"F{r0}")

    # --- Overall progress summary ---
    summary_row = 35
    ws.merge_cells(f"A{summary_row}:H{summary_row}")
    style_section_header(ws[f"A{summary_row}"], "\U0001F4CA OVERALL PROGRESS")

    vals_row = summary_row + 1
    ws[f"A{vals_row}"] = "Total Saved"
    ws[f"A{vals_row}"].font = Font(name=FONT_NAME, bold=True)
    ws[f"B{vals_row}"] = "=" + "+".join(saved_cells)
    style_currency(ws[f"B{vals_row}"])

    ws[f"C{vals_row}"] = "Total Target"
    ws[f"C{vals_row}"].font = Font(name=FONT_NAME, bold=True)
    ws[f"D{vals_row}"] = "=" + "+".join(target_cells)
    style_currency(ws[f"D{vals_row}"])

    ws[f"E{vals_row}"] = "Total Remaining"
    ws[f"E{vals_row}"].font = Font(name=FONT_NAME, bold=True)
    ws[f"F{vals_row}"] = f"=D{vals_row}-B{vals_row}"
    style_currency(ws[f"F{vals_row}"])

    ws[f"G{vals_row}"] = "Overall %"
    ws[f"G{vals_row}"].font = Font(name=FONT_NAME, bold=True)
    ws[f"H{vals_row}"] = f"=IFERROR(B{vals_row}/D{vals_row},0)"
    style_percent(ws[f"H{vals_row}"])

    for col in "ABCDEFGH":
        apply_border(ws[f"{col}{vals_row}"])
        ws[f"{col}{vals_row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    bar_row = vals_row + 1
    ws.merge_cells(f"A{bar_row}:H{bar_row}")
    overall_bar = ws[f"A{bar_row}"]
    overall_bar.value = progress_bar_expr(f"IFERROR(B{vals_row}/D{vals_row},0)")
    overall_bar.font = Font(name=FONT_NAME, size=13, bold=True, color=POSITIVE_TEXT)
    overall_bar.alignment = Alignment(horizontal="left", vertical="center")
    apply_border(overall_bar)

    wb.defined_names.add(DefinedName("GoalData", attr_text="'Savings Goals'!$A$4:$H$33"))

    add_back_link(ws)
    ws.freeze_panes = "A4"
    finalize_sheet(ws)
    return ws


def build_debt_tracker(wb):
    print("✅ Building Debt Tracker...")
    ws = wb.create_sheet("Debt Tracker")
    ws.sheet_properties.tabColor = "D4756B"
    set_col_widths(ws, {
        "A": 20, "B": 16, "C": 14, "D": 14, "E": 10, "F": 13, "G": 12,
        "H": 13, "I": 13, "J": 14, "K": 14, "L": 10, "M": 24, "N": 14,
    })

    ws.merge_cells("A1:N1")
    style_title(ws["A1"], "\U0001F4B3 Debt Tracker")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:N2")
    ws["A2"] = ("List every debt you're paying down. Enter balances, rates, and payments "
                 "in the yellow cells — payoff dates, progress bars, and payoff strategies "
                 "update automatically.")
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=DARK_GRAY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = [
        "Debt Name", "Debt Type", "Starting Balance", "Current Balance",
        "Interest Rate", "Min. Payment", "Extra Payment", "Total Payment",
        "Monthly Interest", "Principal Paid", "Est. Payoff Date", "Progress %",
        "Progress", "Payoff Date (raw)",
    ]
    header_row = 3
    for col, h in zip("ABCDEFGHIJKLMN", headers):
        style_col_header(ws[f"{col}{header_row}"], h)

    data_first = 4
    examples = [
        ("Credit Card A", "Credit Card", 8000, 5400, 0.2299, 150, 100),
        ("Car Loan", "Car Loan", 22000, 14500, 0.069, 380, 0),
        ("Student Loan", "Student Loan", 30000, 24800, 0.05, 320, 0),
        ("Personal Loan", "Personal Loan", 6000, 3100, 0.12, 180, 50),
    ]
    data_last = data_first + len(examples) - 1

    for i, (name, dtype, start_bal, cur_bal, rate, min_pay, extra) in enumerate(examples):
        row = data_first + i
        ws[f"A{row}"] = name
        ws[f"B{row}"] = dtype
        ws[f"C{row}"] = start_bal
        ws[f"D{row}"] = cur_bal
        ws[f"E{row}"] = rate
        ws[f"F{row}"] = min_pay
        ws[f"G{row}"] = extra

    for row in range(data_first, data_last + 1):
        for col in "CD":
            style_currency(ws[f"{col}{row}"])
        ws[f"E{row}"].number_format = "0.00%"
        for col in "FG":
            style_currency(ws[f"{col}{row}"])

        ws[f"H{row}"] = f"=F{row}+G{row}"
        style_currency(ws[f"H{row}"])
        ws[f"I{row}"] = f"=D{row}*(E{row}/12)"
        style_currency(ws[f"I{row}"])
        ws[f"J{row}"] = f"=H{row}-I{row}"
        style_currency(ws[f"J{row}"])
        ws[f"N{row}"] = (
            f'=IF(D{row}<=0,TODAY(),IF(J{row}<=0,DATE(2099,12,31),'
            f'EDATE(TODAY(),ROUND(D{row}/J{row},0))))'
        )
        style_date_cell(ws[f"N{row}"])
        ws[f"K{row}"] = (
            f'=IF(D{row}<=0,"✅ Paid Off!",IF(J{row}<=0,'
            f'"⚠️ Add extra payment",TEXT(N{row},"YYYY-MM-DD")))'
        )
        ws[f"L{row}"] = f"=IFERROR((C{row}-D{row})/C{row},0)"
        style_percent(ws[f"L{row}"])
        ws[f"M{row}"] = progress_bar_expr(f"IFERROR((C{row}-D{row})/C{row},0)")
        ws[f"M{row}"].font = Font(name=FONT_NAME, size=10, color=POSITIVE_TEXT)

        for col in "ABCDEFGHIJKLMN":
            apply_border(ws[f"{col}{row}"])
            if col in "ABCDEFG":
                ws[f"{col}{row}"].protection = Protection(locked=False)
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=YELLOW_INPUT)

    ws.column_dimensions["N"].hidden = True

    alternate_row_colors(ws, data_first, data_last, 1, 14)

    add_dropdown(ws, f"B{data_first}:B{data_last}", "DEBT_TYPES")

    # --- Conditional formatting ---
    ws.conditional_formatting.add(
        f"A{data_first}:M{data_last}",
        FormulaRule(formula=[f"$D{data_first}<=0"],
                    fill=PatternFill("solid", fgColor=LIGHT_GREEN)),
    )
    ws.conditional_formatting.add(
        f"E{data_first}:E{data_last}",
        CellIsRule(operator="greaterThan", formula=["0.2"],
                   fill=PatternFill("solid", fgColor=LIGHT_RED),
                   font=Font(name=FONT_NAME, color=SOFT_RED, bold=True)),
    )

    # --- Debt summary ---
    summary_row = data_last + 2
    ws.merge_cells(f"A{summary_row}:N{summary_row}")
    style_section_header(ws[f"A{summary_row}"], "\U0001F4CA DEBT SUMMARY")

    vals_row = summary_row + 1
    kpis = [
        ("A", "Total Debt", f"=SUM(D{data_first}:D{data_last})", "currency"),
        ("C", "Total Min. Payments", f"=SUM(F{data_first}:F{data_last})", "currency"),
        ("E", "Total Monthly Payment", f"=SUM(H{data_first}:H{data_last})", "currency"),
        ("G", "Total Monthly Interest", f"=SUM(I{data_first}:I{data_last})", "currency"),
        ("I", "Debt-Free Date", f"=TEXT(MAX(N{data_first}:N{data_last}),\"YYYY-MM-DD\")", "text"),
        ("K", "Avg. Interest Rate", f"=AVERAGE(E{data_first}:E{data_last})", "percent"),
    ]
    for label_col, label, formula, kind in kpis:
        value_col = get_column_letter(ord(label_col) - 64 + 1)
        lc = ws[f"{label_col}{vals_row}"]
        lc.value = label
        lc.font = Font(name=FONT_NAME, bold=True, color=DARK_GRAY)
        lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        vc = ws[f"{value_col}{vals_row}"]
        vc.value = formula
        if kind == "currency":
            style_currency(vc)
        elif kind == "percent":
            style_percent(vc)

    ws.row_dimensions[vals_row].height = 30
    for col in "ABCDEFGHIJKLMN":
        apply_border(ws[f"{col}{vals_row}"])
        ws[f"{col}{vals_row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    # --- Avalanche vs. Snowball strategy tables ---
    strategy_header_row = vals_row + 2
    ws.merge_cells(f"A{strategy_header_row}:D{strategy_header_row}")
    style_section_header(ws[f"A{strategy_header_row}"], "\U0001F3D4️ AVALANCHE METHOD — Highest Interest First")
    ws.merge_cells(f"H{strategy_header_row}:K{strategy_header_row}")
    style_section_header(ws[f"H{strategy_header_row}"], "⚪ SNOWBALL METHOD — Smallest Balance First")

    table_header_row = strategy_header_row + 1
    for col, h in zip("ABCD", ["Rank", "Debt Name", "Current Balance", "Interest Rate"]):
        style_col_header(ws[f"{col}{table_header_row}"], h)
    for col, h in zip("HIJK", ["Rank", "Debt Name", "Current Balance", "Monthly Payment"]):
        style_col_header(ws[f"{col}{table_header_row}"], h)

    n_debts = data_last - data_first + 1
    for rank in range(1, n_debts + 1):
        row = table_header_row + rank
        ws[f"A{row}"] = rank
        ws[f"A{row}"].alignment = Alignment(horizontal="center")
        ws[f"D{row}"] = f"=LARGE($E${data_first}:$E${data_last},{rank})"
        ws[f"B{row}"] = (f'=IFERROR(INDEX($A${data_first}:$A${data_last},'
                         f'MATCH(D{row},$E${data_first}:$E${data_last},0)),"")')
        ws[f"C{row}"] = (f'=IFERROR(INDEX($D${data_first}:$D${data_last},'
                         f'MATCH(D{row},$E${data_first}:$E${data_last},0)),"")')
        style_currency(ws[f"C{row}"])
        ws[f"D{row}"].number_format = "0.00%"

        ws[f"H{row}"] = rank
        ws[f"H{row}"].alignment = Alignment(horizontal="center")
        ws[f"J{row}"] = f"=SMALL($D${data_first}:$D${data_last},{rank})"
        ws[f"I{row}"] = (f'=IFERROR(INDEX($A${data_first}:$A${data_last},'
                         f'MATCH(J{row},$D${data_first}:$D${data_last},0)),"")')
        ws[f"K{row}"] = (f'=IFERROR(INDEX($H${data_first}:$H${data_last},'
                         f'MATCH(J{row},$D${data_first}:$D${data_last},0)),"")')
        style_currency(ws[f"J{row}"])
        style_currency(ws[f"K{row}"])

        for col in "ABCD":
            apply_border(ws[f"{col}{row}"])
        for col in "HIJK":
            apply_border(ws[f"{col}{row}"])
        if rank % 2 == 0:
            for col in "ABCD":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)
            for col in "HIJK":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    note_row = table_header_row + n_debts + 2
    ws.merge_cells(f"A{note_row}:N{note_row}")
    note_cell = ws[f"A{note_row}"]
    note_cell.value = (
        "\U0001F4A1 Avalanche saves the most money on interest over time. Snowball builds "
        "momentum with quick wins. Either way: keep paying the minimum on every debt, and "
        "put every extra dollar toward the #1 debt in your chosen method."
    )
    note_cell.font = Font(name=FONT_NAME, size=10, italic=True, color=DARK_GRAY)
    note_cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[note_row].height = 30

    wb.defined_names.add(DefinedName(
        "DebtData", attr_text=f"'Debt Tracker'!$A${data_first}:$M${data_last}"))

    add_back_link(ws)
    ws.freeze_panes = "B4"
    finalize_sheet(ws)
    return ws


def build_sinking_funds(wb):
    print("✅ Building Sinking Funds...")
    ws = wb.create_sheet("Sinking Funds")
    ws.sheet_properties.tabColor = "B2D8D8"
    set_col_widths(ws, {
        "A": 22, "B": 16, "C": 15, "D": 15, "E": 15, "F": 15, "G": 11, "H": 24, "I": 14,
    })

    ws.merge_cells("A1:I1")
    style_title(ws["A1"], "\U0001FAD9 Sinking Funds")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:I2")
    ws["A2"] = ("Sinking funds are savings set aside for known upcoming expenses (car "
                 "repairs, holiday gifts, annual bills). Enter your annual target and "
                 "current balance for each fund — the rest updates automatically.")
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=DARK_GRAY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = [
        "Fund Name", "Category", "Annual Target", "Current Balance",
        "Suggested Monthly", "Remaining", "% Funded", "Progress", "Next Use Date",
    ]
    header_row = 3
    for col, h in zip("ABCDEFGHI", headers):
        style_col_header(ws[f"{col}{header_row}"], h)

    data_first = 4
    today = date.today()
    examples = [
        ("Car Maintenance", "Auto", 1200, 450, today + timedelta(days=200)),
        ("Holiday Gifts", "Personal", 800, 800, date(today.year, 12, 25)),
        ("Property Taxes", "Home", 2400, 600, today + timedelta(days=300)),
        ("Annual Subscriptions", "Subscriptions", 300, 300, today + timedelta(days=60)),
        ("Home Maintenance", "Home", 1500, 375, today + timedelta(days=180)),
        ("Medical / Dental", "Health", 1000, 250, today + timedelta(days=150)),
        ("Car Registration", "Auto", 200, 200, today + timedelta(days=60)),
        ("Birthday Gifts", "Personal", 400, 100, today + timedelta(days=90)),
        ("Pet Care", "Pets", 600, 300, today + timedelta(days=120)),
        ("Clothing", "Personal", 600, 150, today + timedelta(days=200)),
        ("Electronics / Tech", "Personal", 500, 500, today + timedelta(days=250)),
        ("Travel Fund", "Travel", 2000, 800, today + timedelta(days=180)),
    ]
    data_last = data_first + len(examples) - 1

    for i, (name, cat, target, balance, next_use) in enumerate(examples):
        row = data_first + i
        ws[f"A{row}"] = name
        ws[f"B{row}"] = cat
        ws[f"C{row}"] = target
        ws[f"D{row}"] = balance
        ws[f"I{row}"] = next_use

    for row in range(data_first, data_last + 1):
        style_currency(ws[f"C{row}"])
        style_currency(ws[f"D{row}"])
        ws[f"E{row}"] = f"=ROUND(C{row}/12,2)"
        style_currency(ws[f"E{row}"])
        ws[f"F{row}"] = f"=C{row}-D{row}"
        style_currency(ws[f"F{row}"])
        ws[f"G{row}"] = f"=IFERROR(D{row}/C{row},0)"
        style_percent(ws[f"G{row}"])
        ws[f"H{row}"] = progress_bar_expr(f"IFERROR(MIN(D{row}/C{row},1),0)")
        ws[f"H{row}"].font = Font(name=FONT_NAME, size=10, color=POSITIVE_TEXT)
        style_date_cell(ws[f"I{row}"])

        for col in "ABCDEFGHI":
            apply_border(ws[f"{col}{row}"])
            if col in "ABCDI":
                ws[f"{col}{row}"].protection = Protection(locked=False)
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=YELLOW_INPUT)

    alternate_row_colors(ws, data_first, data_last, 1, 9)

    # --- Conditional formatting on % Funded ---
    pct_range = f"G{data_first}:G{data_last}"
    ws.conditional_formatting.add(pct_range, CellIsRule(
        operator="greaterThanOrEqual", formula=["1"],
        fill=PatternFill("solid", fgColor=LIGHT_GREEN)))
    ws.conditional_formatting.add(pct_range, CellIsRule(
        operator="between", formula=["0.5", "0.999999"],
        fill=PatternFill("solid", fgColor=GOLD)))
    ws.conditional_formatting.add(pct_range, CellIsRule(
        operator="lessThan", formula=["0.5"],
        fill=PatternFill("solid", fgColor=LIGHT_BLUE)))

    # --- Totals row ---
    total_row = data_last + 2
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"A{total_row}"].font = Font(name=FONT_NAME, bold=True, size=12)
    for col, formula in [
        ("C", f"=SUM(C{data_first}:C{data_last})"),
        ("D", f"=SUM(D{data_first}:D{data_last})"),
        ("E", f"=SUM(E{data_first}:E{data_last})"),
        ("F", f"=SUM(F{data_first}:F{data_last})"),
    ]:
        ws[f"{col}{total_row}"] = formula
        style_currency(ws[f"{col}{total_row}"])
        ws[f"{col}{total_row}"].font = Font(name=FONT_NAME, bold=True)

    ws[f"G{total_row}"] = f"=IFERROR(D{total_row}/C{total_row},0)"
    style_percent(ws[f"G{total_row}"])
    ws[f"G{total_row}"].font = Font(name=FONT_NAME, bold=True)

    ws[f"H{total_row}"] = progress_bar_expr(f"IFERROR(D{total_row}/C{total_row},0)")
    ws[f"H{total_row}"].font = Font(name=FONT_NAME, size=10, bold=True, color=POSITIVE_TEXT)

    for col in "ABCDEFGHI":
        apply_border(ws[f"{col}{total_row}"])
        ws[f"{col}{total_row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    wb.defined_names.add(DefinedName(
        "SinkingFundData", attr_text=f"'Sinking Funds'!$A${data_first}:$I${data_last}"))

    add_back_link(ws)
    ws.freeze_panes = "B4"
    finalize_sheet(ws)
    return ws


def build_subscriptions(wb):
    print("✅ Building Subscriptions...")
    ws = wb.create_sheet("Subscriptions")
    ws.sheet_properties.tabColor = "E8D5D5"
    set_col_widths(ws, {
        "A": 22, "B": 16, "C": 13, "D": 13, "E": 13, "F": 16, "G": 16, "H": 11,
    })

    ws.merge_cells("A1:H1")
    style_title(ws["A1"], "\U0001F501 Subscriptions")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = ("Track every recurring subscription in one place. Set a subscription "
                 "to 'No' under Active when you cancel it — totals below only count "
                 "active subscriptions.")
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=DARK_GRAY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = [
        "Service Name", "Category", "Monthly Cost", "Billing Cycle",
        "Annual Cost", "Next Billing Date", "Payment Method", "Active?",
    ]
    header_row = 3
    for col, h in zip("ABCDEFGH", headers):
        style_col_header(ws[f"{col}{header_row}"], h)

    data_first = 4
    today = date.today()
    examples = [
        ("Netflix", "Streaming", 15.49, "Monthly", today + timedelta(days=12), "Credit Card", "Yes"),
        ("Spotify", "Streaming", 11.99, "Monthly", today + timedelta(days=5), "Credit Card", "Yes"),
        ("Amazon Prime", "Shopping", 139.00, "Annual", today + timedelta(days=200), "Credit Card", "Yes"),
        ("Disney+", "Streaming", 13.99, "Monthly", today + timedelta(days=18), "Debit Card", "Yes"),
        ("Gym Membership", "Fitness", 39.99, "Monthly", today + timedelta(days=1), "Bank Transfer", "Yes"),
        ("iCloud Storage", "Software", 2.99, "Monthly", today + timedelta(days=9), "Credit Card", "Yes"),
        ("YouTube Premium", "Streaming", 13.99, "Monthly", today + timedelta(days=22), "Credit Card", "Yes"),
        ("Adobe Creative Cloud", "Software", 54.99, "Monthly", today + timedelta(days=14), "Credit Card", "Yes"),
        ("Microsoft 365", "Software", 99.00, "Annual", today + timedelta(days=120), "Credit Card", "Yes"),
        ("Hulu", "Streaming", 7.99, "Monthly", today + timedelta(days=27), "Debit Card", "No"),
        ("Audible", "Entertainment", 14.95, "Monthly", today + timedelta(days=8), "Credit Card", "Yes"),
        ("Domain / Hosting", "Software", 180.00, "Annual", today + timedelta(days=250), "PayPal", "Yes"),
    ]
    data_last = data_first + len(examples) - 1

    for i, (name, cat, cost, cycle, next_bill, method, active) in enumerate(examples):
        row = data_first + i
        ws[f"A{row}"] = name
        ws[f"B{row}"] = cat
        ws[f"C{row}"] = cost
        ws[f"D{row}"] = cycle
        ws[f"F{row}"] = next_bill
        ws[f"G{row}"] = method
        ws[f"H{row}"] = active

    for row in range(data_first, data_last + 1):
        style_currency(ws[f"C{row}"])
        ws[f"E{row}"] = (
            f'=IFS(D{row}="Monthly",C{row}*12,D{row}="Annual",C{row},'
            f'D{row}="Quarterly",C{row}*4,D{row}="Weekly",C{row}*52,TRUE,C{row}*12)'
        )
        style_currency(ws[f"E{row}"])
        style_date_cell(ws[f"F{row}"])

        for col in "ABCDEFGH":
            apply_border(ws[f"{col}{row}"])
            if col in "ABCDFGH":
                ws[f"{col}{row}"].protection = Protection(locked=False)
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=YELLOW_INPUT)

    alternate_row_colors(ws, data_first, data_last, 1, 8)

    add_dropdown(ws, f"D{data_first}:D{data_last}", "BILLING_CYCLES")
    add_dropdown(ws, f"G{data_first}:G{data_last}", "PAYMENT_METHODS")
    add_dropdown(ws, f"H{data_first}:H{data_last}", "YES_NO")

    # --- Gray out cancelled subscriptions ---
    ws.conditional_formatting.add(
        f"A{data_first}:H{data_last}",
        FormulaRule(formula=[f'$H{data_first}="No"'],
                    fill=PatternFill("solid", fgColor=MEDIUM_GRAY),
                    font=Font(name=FONT_NAME, color=DARK_GRAY, italic=True, strike=True)),
    )

    # --- Summary ---
    summary_row = data_last + 2
    ws.merge_cells(f"A{summary_row}:H{summary_row}")
    style_section_header(ws[f"A{summary_row}"], "\U0001F4CA SUBSCRIPTION SUMMARY")

    vals_row = summary_row + 1
    kpis = [
        ("A", "Active Subscriptions", f'=COUNTIF(H{data_first}:H{data_last},"Yes")', "int"),
        ("C", "Monthly Cost (Active)", f'=SUMIF(H{data_first}:H{data_last},"Yes",C{data_first}:C{data_last})', "currency"),
        ("E", "Annual Cost (Active)", f'=SUMIF(H{data_first}:H{data_last},"Yes",E{data_first}:E{data_last})', "currency"),
        ("G", "Cancelled (Annual Savings)", f'=SUMIF(H{data_first}:H{data_last},"No",E{data_first}:E{data_last})', "currency"),
    ]
    for label_col, label, formula, kind in kpis:
        value_col = get_column_letter(ord(label_col) - 64 + 1)
        lc = ws[f"{label_col}{vals_row}"]
        lc.value = label
        lc.font = Font(name=FONT_NAME, bold=True, color=DARK_GRAY)
        lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        vc = ws[f"{value_col}{vals_row}"]
        vc.value = formula
        if kind == "currency":
            style_currency(vc)
        elif kind == "int":
            vc.number_format = INT_FMT

    ws.row_dimensions[vals_row].height = 30
    for col in "ABCDEFGH":
        apply_border(ws[f"{col}{vals_row}"])
        ws[f"{col}{vals_row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    wb.defined_names.add(DefinedName(
        "SubscriptionData", attr_text=f"'Subscriptions'!$A${data_first}:$H${data_last}"))

    add_back_link(ws)
    ws.freeze_panes = "B4"
    finalize_sheet(ws)
    return ws


def build_bills_calendar(wb):
    print("✅ Building Bills Calendar...")
    ws = wb.create_sheet("Bills Calendar")
    ws.sheet_properties.tabColor = GOLD
    set_col_widths(ws, {
        "A": 22, "B": 15, "C": 9, "D": 12, "E": 15, "F": 10, "G": 9,
        "H": 13, "I": 13, "J": 18, "K": 13, "L": 16,
    })

    ws.merge_cells("A1:L1")
    style_title(ws["A1"], "\U0001F4C5 Bills Calendar")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:L2")
    ws["A2"] = ("List every recurring bill with its due day of the month. Mark a bill "
                 "'Yes' under Paid? once you've paid it — the Status column and summary "
                 "below update automatically.")
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=DARK_GRAY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    headers = [
        "Bill Name", "Category", "Due Day", "Amount", "Payment Method", "Auto-Pay?",
        "Paid?", "Due Date", "Days Until Due", "Notes", "Date Paid", "Status",
    ]
    header_row = 3
    for col, h in zip("ABCDEFGHIJKL", headers):
        style_col_header(ws[f"{col}{header_row}"], h)

    data_first = 4
    today = date.today()
    examples = [
        ("Rent / Mortgage", "Housing", 1, 1450, "Bank Transfer", "Yes", "Yes"),
        ("Electricity", "Utilities", 5, 120, "Debit Card", "Yes", "Yes"),
        ("Water", "Utilities", 7, 45, "Debit Card", "No", "No"),
        ("Internet", "Utilities", 10, 65, "Credit Card", "Yes", "Yes"),
        ("Phone", "Utilities", 12, 85, "Credit Card", "Yes", "No"),
        ("Car Insurance", "Insurance", 15, 110, "Bank Transfer", "Yes", "No"),
        ("Health Insurance", "Insurance", 1, 320, "Bank Transfer", "Yes", "Yes"),
        ("Car Payment", "Debt Payment", 20, 380, "Bank Transfer", "Yes", "No"),
        ("Student Loan Payment", "Debt Payment", 25, 320, "Bank Transfer", "Yes", "No"),
        ("Credit Card Payment", "Debt Payment", 28, 150, "Bank Transfer", "No", "No"),
        ("Streaming Bundle", "Subscriptions", 3, 45, "Credit Card", "Yes", "Yes"),
        ("Gym Membership", "Personal Care", 1, 40, "Bank Transfer", "Yes", "Yes"),
        ("Trash / Recycling", "Utilities", 18, 25, "Debit Card", "No", "No"),
        ("Home / Renters Insurance", "Insurance", 1, 35, "Bank Transfer", "Yes", "Yes"),
        ("Childcare", "Kids", 1, 850, "Bank Transfer", "No", "Yes"),
    ]
    data_last = data_first + len(examples) - 1

    for i, (name, cat, due_day, amount, method, autopay, paid) in enumerate(examples):
        row = data_first + i
        ws[f"A{row}"] = name
        ws[f"B{row}"] = cat
        ws[f"C{row}"] = due_day
        ws[f"D{row}"] = amount
        ws[f"E{row}"] = method
        ws[f"F{row}"] = autopay
        ws[f"G{row}"] = paid
        if paid == "Yes":
            ws[f"K{row}"] = today - timedelta(days=2)

    for row in range(data_first, data_last + 1):
        ws[f"C{row}"].alignment = Alignment(horizontal="center")
        style_currency(ws[f"D{row}"])
        ws[f"H{row}"] = f"=DATE(YEAR(TODAY()),MONTH(TODAY()),C{row})"
        style_date_cell(ws[f"H{row}"])
        ws[f"I{row}"] = f"=H{row}-TODAY()"
        ws[f"I{row}"].number_format = INT_FMT
        ws[f"I{row}"].alignment = Alignment(horizontal="center")
        if ws[f"K{row}"].value is not None:
            style_date_cell(ws[f"K{row}"])
        ws[f"L{row}"] = (
            f'=IF(G{row}="Yes","✅ Paid",IF(I{row}<0,"\U0001F534 Overdue",'
            f'IF(I{row}<=3,"⚠️ Due Soon","\U0001F4C5 Upcoming")))'
        )

        for col in "ABCDEFGHIJKL":
            apply_border(ws[f"{col}{row}"])
            if col in "ABCDEFGJK":
                ws[f"{col}{row}"].protection = Protection(locked=False)
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=YELLOW_INPUT)

    alternate_row_colors(ws, data_first, data_last, 1, 12)

    add_dropdown(ws, f"E{data_first}:E{data_last}", "PAYMENT_METHODS")
    add_dropdown(ws, f"F{data_first}:F{data_last}", "YES_NO")
    add_dropdown(ws, f"G{data_first}:G{data_last}", "YES_NO")

    # --- Status conditional formatting ---
    status_range = f"L{data_first}:L{data_last}"
    ws.conditional_formatting.add(status_range, CellIsRule(
        operator="equal", formula=['"✅ Paid"'],
        fill=PatternFill("solid", fgColor=LIGHT_GREEN)))
    ws.conditional_formatting.add(status_range, CellIsRule(
        operator="equal", formula=['"\U0001F534 Overdue"'],
        fill=PatternFill("solid", fgColor=LIGHT_RED),
        font=Font(name=FONT_NAME, color=SOFT_RED, bold=True)))
    ws.conditional_formatting.add(status_range, CellIsRule(
        operator="equal", formula=['"⚠️ Due Soon"'],
        fill=PatternFill("solid", fgColor=GOLD)))
    ws.conditional_formatting.add(status_range, CellIsRule(
        operator="equal", formula=['"\U0001F4C5 Upcoming"'],
        fill=PatternFill("solid", fgColor=LIGHT_BLUE)))

    # --- Summary ---
    summary_row = data_last + 2
    ws.merge_cells(f"A{summary_row}:L{summary_row}")
    style_section_header(ws[f"A{summary_row}"], "\U0001F4CA BILLS SUMMARY")

    vals_row = summary_row + 1
    kpis = [
        ("A", "Total Monthly Bills", f"=SUM(D{data_first}:D{data_last})", "currency"),
        ("C", "Bills Paid", f'=COUNTIF(G{data_first}:G{data_last},"Yes")', "int"),
        ("E", "Bills Remaining", f'=COUNTIF(G{data_first}:G{data_last},"No")', "int"),
        ("G", "Overdue", f'=COUNTIF(L{data_first}:L{data_last},"\U0001F534 Overdue")', "int"),
        ("I", "Due Soon (≤3 days)", f'=COUNTIF(L{data_first}:L{data_last},"⚠️ Due Soon")', "int"),
    ]
    for label_col, label, formula, kind in kpis:
        value_col = get_column_letter(ord(label_col) - 64 + 1)
        lc = ws[f"{label_col}{vals_row}"]
        lc.value = label
        lc.font = Font(name=FONT_NAME, bold=True, color=DARK_GRAY)
        lc.alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
        vc = ws[f"{value_col}{vals_row}"]
        vc.value = formula
        if kind == "currency":
            style_currency(vc)
        elif kind == "int":
            vc.number_format = INT_FMT

    ws.row_dimensions[vals_row].height = 30
    for col in "ABCDEFGHIJKL":
        apply_border(ws[f"{col}{vals_row}"])
        ws[f"{col}{vals_row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    wb.defined_names.add(DefinedName(
        "BillsData", attr_text=f"'Bills Calendar'!$A${data_first}:$L${data_last}"))

    add_back_link(ws)
    ws.freeze_panes = "B4"
    finalize_sheet(ws)
    return ws


def build_annual_overview(wb):
    print("✅ Building Annual Overview...")
    ws = wb.create_sheet("Annual Overview")
    ws.sheet_properties.tabColor = ACCENT_BLUE
    set_col_widths(ws, {"A": 16, "B": 15, "C": 15, "D": 15, "E": 14, "F": 28})

    ws.merge_cells("A1:F1")
    style_title(ws["A1"], "\U0001F4C8 Annual Overview")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:F2")
    ws["A2"] = ("A year-at-a-glance view of your income, expenses, and savings — "
                 "built automatically from your Income Tracker and Expense Tracker.")
    ws["A2"].font = Font(name=FONT_NAME, size=10, italic=True, color=DARK_GRAY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Layout constants
    monthly_header_row = 7
    monthly_colheader_row = 8
    monthly_data_first = 9
    monthly_data_last = 20
    monthly_total_row = 21
    cat_header_row = 23
    cat_colheader_row = 24
    cat_data_first = 25
    cat_data_last = cat_data_first + len(CATEGORIES_LIST) - 1

    # --- Year-to-date KPI section ---
    kpi_header_row = 3
    ws.merge_cells(f"A{kpi_header_row}:F{kpi_header_row}")
    style_section_header(ws[f"A{kpi_header_row}"],
                          '="\U0001F4CA YEAR-TO-DATE SUMMARY — "&TEXT(TODAY(),"YYYY")')

    kpi_label_row = 4
    kpi_value_row = 5
    kpi_labels = [
        "Total Income (YTD)", "Total Expenses (YTD)", "Net Savings (YTD)",
        "Savings Rate (YTD)", "Avg Monthly Income", "Avg Monthly Expenses",
    ]
    for col, label in zip("ABCDEF", kpi_labels):
        style_kpi_label(ws[f"{col}{kpi_label_row}"], label)

    kpi_formulas = {
        "A": (f"=SUM(B{monthly_data_first}:B{monthly_data_last})", "currency"),
        "B": (f"=SUM(C{monthly_data_first}:C{monthly_data_last})", "currency"),
        "C": (f"=A{kpi_value_row}-B{kpi_value_row}", "currency"),
        "D": (f"=IFERROR(C{kpi_value_row}/A{kpi_value_row},0)", "percent"),
        "E": (f"=AVERAGE(B{monthly_data_first}:B{monthly_data_last})", "currency"),
        "F": (f"=AVERAGE(C{monthly_data_first}:C{monthly_data_last})", "currency"),
    }
    for col, (formula, kind) in kpi_formulas.items():
        cell = ws[f"{col}{kpi_value_row}"]
        cell.value = formula
        cell.font = Font(name=FONT_NAME, size=14, bold=True, color=DARK_TEXT)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        if kind == "currency":
            style_currency(cell)
        else:
            style_percent(cell)
        apply_border(cell)
    ws.row_dimensions[kpi_value_row].height = 24

    # --- Monthly breakdown table ---
    ws.merge_cells(f"A{monthly_header_row}:F{monthly_header_row}")
    style_section_header(ws[f"A{monthly_header_row}"], "\U0001F4C5 MONTHLY BREAKDOWN")

    for col, h in zip("ABCDEF", ["Month", "Income", "Expenses", "Net", "Savings Rate", "% of Income Spent"]):
        style_col_header(ws[f"{col}{monthly_colheader_row}"], h)

    for m in range(1, 13):
        row = monthly_data_first + m - 1
        ws[f"A{row}"] = date(2000, m, 1).strftime("%B")
        ws[f"B{row}"] = (f"=SUMIFS('Income Tracker'!G4:G103,"
                         f"'Income Tracker'!J4:J103,{m},"
                         f"'Income Tracker'!K4:K103,YEAR(TODAY()))")
        ws[f"C{row}"] = (f"=SUMIFS('Expense Tracker'!E4:E503,"
                         f"'Expense Tracker'!L4:L503,{m},"
                         f"'Expense Tracker'!M4:M503,YEAR(TODAY()))")
        ws[f"D{row}"] = f"=B{row}-C{row}"
        ws[f"E{row}"] = f"=IFERROR(D{row}/B{row},0)"
        ws[f"F{row}"] = progress_bar_expr(f"IFERROR(MIN(C{row}/B{row},1),0)")
        ws[f"F{row}"].font = Font(name=FONT_NAME, size=10, color=POSITIVE_TEXT)
        for col in "BCD":
            style_currency(ws[f"{col}{row}"])
        style_percent(ws[f"E{row}"])
        for col in "ABCDEF":
            apply_border(ws[f"{col}{row}"])
        if m % 2 == 0:
            for col in "ABCDEF":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    ws[f"A{monthly_total_row}"] = "TOTAL"
    ws[f"A{monthly_total_row}"].font = Font(name=FONT_NAME, bold=True, size=12)
    ws[f"B{monthly_total_row}"] = f"=SUM(B{monthly_data_first}:B{monthly_data_last})"
    ws[f"C{monthly_total_row}"] = f"=SUM(C{monthly_data_first}:C{monthly_data_last})"
    ws[f"D{monthly_total_row}"] = f"=B{monthly_total_row}-C{monthly_total_row}"
    ws[f"E{monthly_total_row}"] = f"=IFERROR(D{monthly_total_row}/B{monthly_total_row},0)"
    ws[f"F{monthly_total_row}"] = progress_bar_expr(
        f"IFERROR(MIN(C{monthly_total_row}/B{monthly_total_row},1),0)")
    ws[f"F{monthly_total_row}"].font = Font(name=FONT_NAME, size=10, bold=True, color=POSITIVE_TEXT)
    for col in "BCD":
        style_currency(ws[f"{col}{monthly_total_row}"])
    style_percent(ws[f"E{monthly_total_row}"])
    for col in "ABCDEF":
        apply_border(ws[f"{col}{monthly_total_row}"])
        ws[f"{col}{monthly_total_row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)
        ws[f"{col}{monthly_total_row}"].font = Font(name=FONT_NAME, bold=True)

    cf_diff_green_red(ws, f"D{monthly_data_first}:D{monthly_total_row}")

    savings_rate_range = f"E{monthly_data_first}:E{monthly_total_row}"
    ws.conditional_formatting.add(savings_rate_range, CellIsRule(
        operator="greaterThanOrEqual", formula=["0.2"],
        fill=PatternFill("solid", fgColor=LIGHT_GREEN)))
    ws.conditional_formatting.add(savings_rate_range, CellIsRule(
        operator="between", formula=["0.1", "0.199999"],
        fill=PatternFill("solid", fgColor=GOLD)))
    ws.conditional_formatting.add(savings_rate_range, CellIsRule(
        operator="lessThan", formula=["0.1"],
        fill=PatternFill("solid", fgColor=LIGHT_RED)))

    # --- Spending by category (year) ---
    ws.merge_cells(f"A{cat_header_row}:F{cat_header_row}")
    style_section_header(ws[f"A{cat_header_row}"], "\U0001F4CA SPENDING BY CATEGORY (YEAR)")

    for col, h in zip("ABCDEF", ["Category", "Annual Total", "% of Annual Spending",
                                  "Avg Monthly", "", "Relative Size"]):
        style_col_header(ws[f"{col}{cat_colheader_row}"], h)
    ws.merge_cells(f"E{cat_colheader_row}:F{cat_colheader_row}")

    for i, cat in enumerate(CATEGORIES_LIST):
        row = cat_data_first + i
        ws[f"A{row}"] = cat
        ws[f"B{row}"] = (f"=SUMIFS('Expense Tracker'!E4:E503,"
                         f"'Expense Tracker'!C4:C503,A{row},"
                         f"'Expense Tracker'!M4:M503,YEAR(TODAY()))")
        ws[f"C{row}"] = f"=IFERROR(B{row}/$C${monthly_total_row},0)"
        ws[f"D{row}"] = f"=B{row}/12"
        ws.merge_cells(f"E{row}:F{row}")
        ws[f"E{row}"] = progress_bar_expr(
            f"IFERROR(B{row}/MAX($B${cat_data_first}:$B${cat_data_last}),0)")
        ws[f"E{row}"].font = Font(name=FONT_NAME, size=10, color=POSITIVE_TEXT)
        style_currency(ws[f"B{row}"])
        style_percent(ws[f"C{row}"])
        style_currency(ws[f"D{row}"])
        for col in "ABCDEF":
            apply_border(ws[f"{col}{row}"])
        if i % 2 == 1:
            for col in "ABCDEF":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    top5 = Rule(type="top10", rank=5,
                 dxf=DifferentialStyle(fill=PatternFill("solid", fgColor=GOLD)))
    ws.conditional_formatting.add(f"B{cat_data_first}:B{cat_data_last}", top5)

    add_back_link(ws)
    ws.freeze_panes = "A3"
    finalize_sheet(ws)
    return ws


def build_dashboard(wb):
    print("✅ Building Dashboard...")
    ws = wb.create_sheet("Dashboard")
    ws.sheet_properties.tabColor = PRIMARY_GREEN
    set_col_widths(ws, {
        "A": 18, "B": 16, "C": 16, "D": 18, "E": 18, "F": 16, "G": 16, "H": 18,
        "J": 16, "K": 12, "L": 8, "M": 4, "N": 6, "O": 6,
    })

    ws.merge_cells("A1:H1")
    style_title(ws["A1"], "\U0001F3E0 Dashboard")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:H2")
    ws["A2"] = ('="Welcome back, "&\'START HERE\'!B5&"! Here is your financial '
                 'snapshot for "&\'START HERE\'!B7&"."')
    ws["A2"].font = Font(name=FONT_NAME, size=12, italic=True, color=DARK_GRAY)
    ws["A2"].alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[2].height = 30

    # --- KPI section ---
    ws.merge_cells("A3:H3")
    style_section_header(ws["A3"], '="\U0001F4CA THIS MONTH AT A GLANCE — "&TEXT(TODAY(),"MMMM YYYY")')

    def kpi_card(start_col, label, formula, kind, label_row, value_row):
        end_col = get_column_letter(ord(start_col) - 64 + 1)
        ws.merge_cells(f"{start_col}{label_row}:{end_col}{label_row}")
        lc = ws[f"{start_col}{label_row}"]
        lc.value = label
        lc.font = Font(name=FONT_NAME, size=10, bold=False, color=DARK_TEXT)
        lc.fill = PatternFill("solid", fgColor=ACCENT_BLUE)
        lc.alignment = Alignment(horizontal="center", vertical="center")
        apply_border(lc)

        ws.merge_cells(f"{start_col}{value_row}:{end_col}{value_row}")
        vc = ws[f"{start_col}{value_row}"]
        vc.value = formula
        vc.font = Font(name=FONT_NAME, size=16, bold=True, color=DARK_TEXT)
        vc.alignment = Alignment(horizontal="center", vertical="center")
        if kind == "currency":
            style_currency(vc)
        elif kind == "percent":
            style_percent(vc)
        elif kind == "int":
            vc.number_format = INT_FMT
        apply_border(vc)

    kpi_card("A", "\U0001F4B0 Income This Month",
             "=SUMIFS('Income Tracker'!G4:G103,'Income Tracker'!J4:J103,MONTH(TODAY()),"
             "'Income Tracker'!K4:K103,YEAR(TODAY()))", "currency", 4, 5)
    kpi_card("C", "\U0001F4B8 Expenses This Month",
             "=SUMIFS('Expense Tracker'!E4:E503,'Expense Tracker'!L4:L503,MONTH(TODAY()),"
             "'Expense Tracker'!M4:M503,YEAR(TODAY()))", "currency", 4, 5)
    kpi_card("E", "\U0001F4C8 Net This Month", "=A5-C5", "currency", 4, 5)
    kpi_card("G", "\U0001F3AF Savings Rate", "=IFERROR(E5/A5,0)", "percent", 4, 5)

    kpi_card("A", "\U0001F4B3 Total Debt", "=SUM('Debt Tracker'!D4:D7)", "currency", 6, 7)
    kpi_card("C", "\U0001F3C6 Total Saved (Goals)", "='Savings Goals'!B36", "currency", 6, 7)
    kpi_card("E", "\U0001F4C5 Bills Due This Month", "=SUM('Bills Calendar'!D4:D18)", "currency", 6, 7)
    kpi_card("G", "\U0001F514 Unpaid Bills", '=COUNTIF(\'Bills Calendar\'!G4:G18,"No")', "int", 6, 7)

    ws.row_dimensions[4].height = 20
    ws.row_dimensions[5].height = 32
    ws.row_dimensions[6].height = 20
    ws.row_dimensions[7].height = 32

    cf_diff_green_red(ws, "E5:E5")
    ws.conditional_formatting.add("G5:G5", CellIsRule(
        operator="greaterThanOrEqual", formula=["0.2"],
        fill=PatternFill("solid", fgColor=LIGHT_GREEN)))
    ws.conditional_formatting.add("G5:G5", CellIsRule(
        operator="lessThan", formula=["0.1"],
        fill=PatternFill("solid", fgColor=LIGHT_RED)))

    ws.row_dimensions[8].height = 8  # spacer

    # --- Spending by Category | Savings Goals Progress ---
    panel_row = 9
    ws.merge_cells(f"A{panel_row}:D{panel_row}")
    style_section_header(ws[f"A{panel_row}"], "\U0001F4CA SPENDING BY CATEGORY (THIS MONTH)")
    ws.merge_cells(f"E{panel_row}:H{panel_row}")
    style_section_header(ws[f"E{panel_row}"], "\U0001F3AF SAVINGS GOALS PROGRESS")

    # Hidden helper: per-category spend this month
    helper_first = 4
    for i, cat in enumerate(CATEGORIES_LIST):
        hrow = helper_first + i
        ws[f"J{hrow}"] = cat
        ws[f"K{hrow}"] = (f"=SUMIFS('Expense Tracker'!E4:E503,"
                          f"'Expense Tracker'!C4:C503,J{hrow},"
                          f"'Expense Tracker'!L4:L503,MONTH(TODAY()),"
                          f"'Expense Tracker'!M4:M503,YEAR(TODAY()))")
    helper_last = helper_first + len(CATEGORIES_LIST) - 1

    data_start = panel_row + 1  # 10
    for rank in range(1, 7):
        row = data_start + rank - 1
        ws[f"B{row}"] = f"=LARGE($K${helper_first}:$K${helper_last},{rank})"
        ws[f"A{row}"] = (f'=IF(B{row}<=0,"—",INDEX($J${helper_first}:$J${helper_last},'
                         f'MATCH(B{row},$K${helper_first}:$K${helper_last},0)))')
        style_currency(ws[f"B{row}"])
        ws.merge_cells(f"C{row}:D{row}")
        pct_expr = (f"IFERROR(MIN(B{row}/INDEX('Monthly Budget'!E4:E103,"
                    f"MATCH(A{row},'Monthly Budget'!B4:B103,0)),1),0)")
        ws[f"C{row}"] = progress_bar_expr(pct_expr)
        ws[f"C{row}"].font = Font(name=FONT_NAME, size=10, color=POSITIVE_TEXT)
        for col in "ABCD":
            apply_border(ws[f"{col}{row}"])
        if rank % 2 == 0:
            for col in "ABCD":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    # Savings goals progress (right panel) — 6 goals, one per row
    for i in range(6):
        row = data_start + i
        r0 = 4 + i * 5
        r1 = r0 + 1
        r2 = r0 + 2
        ws[f"E{row}"] = f"='Savings Goals'!A{r0}"
        ws[f"F{row}"] = (f'=TEXT(\'Savings Goals\'!B{r1},"$#,##0")&" / "&'
                         f'TEXT(\'Savings Goals\'!F{r0},"$#,##0")')
        ws.merge_cells(f"G{row}:H{row}")
        ws[f"G{row}"] = f"='Savings Goals'!A{r2}"
        ws[f"G{row}"].font = Font(name=FONT_NAME, size=10, color=POSITIVE_TEXT)
        for col in "EFGH":
            apply_border(ws[f"{col}{row}"])
        if i % 2 == 1:
            for col in "EFGH":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    ws.row_dimensions[16].height = 8  # spacer

    # --- Debt Snapshot | Upcoming Bills ---
    panel_row2 = 17
    ws.merge_cells(f"A{panel_row2}:D{panel_row2}")
    style_section_header(ws[f"A{panel_row2}"], "\U0001F4B3 DEBT SNAPSHOT")
    ws.merge_cells(f"E{panel_row2}:H{panel_row2}")
    style_section_header(ws[f"E{panel_row2}"], "\U0001F4C5 UPCOMING BILLS")

    data_start2 = panel_row2 + 1  # 18

    # Hidden helper: bills sort key (unpaid bills sort by days-until-due; paid -> 9999)
    for k in range(15):
        brow = 4 + k
        ws[f"L{brow}"] = (f'=IF(\'Bills Calendar\'!G{brow}="Yes",9999,'
                          f"'Bills Calendar'!I{brow})")
    # Hidden helper: matched row index for the 4 most urgent unpaid bills
    for k in range(4):
        nrow = 4 + k
        ws[f"N{nrow}"] = (f"=IFERROR(IF(SMALL($L$4:$L$18,{k + 1})>=9999,0,"
                          f"MATCH(SMALL($L$4:$L$18,{k + 1}),$L$4:$L$18,0)),0)")

    for j in range(4):
        row = data_start2 + j
        debt_row = 4 + j
        ws[f"A{row}"] = f"='Debt Tracker'!A{debt_row}"
        ws[f"B{row}"] = f"='Debt Tracker'!D{debt_row}"
        style_currency(ws[f"B{row}"])
        ws.merge_cells(f"C{row}:D{row}")
        ws[f"C{row}"] = f"='Debt Tracker'!M{debt_row}"
        ws[f"C{row}"].font = Font(name=FONT_NAME, size=10, color=POSITIVE_TEXT)

        nref = f"N{4 + j}"
        ws[f"E{row}"] = (
            f'=IF({nref}=0,"\U0001F389 All bills paid!",'
            f'INDEX(\'Bills Calendar\'!$A$4:$A$18,{nref})&" — "&'
            f'TEXT(INDEX(\'Bills Calendar\'!$H$4:$H$18,{nref}),"MMM D"))'
        )
        ws[f"F{row}"] = f'=IF({nref}=0,"",INDEX(\'Bills Calendar\'!$D$4:$D$18,{nref}))'
        style_currency(ws[f"F{row}"])
        ws.merge_cells(f"G{row}:H{row}")
        ws[f"G{row}"] = f'=IF({nref}=0,"",INDEX(\'Bills Calendar\'!$L$4:$L$18,{nref}))'

        for col in "ABCDEFGH":
            apply_border(ws[f"{col}{row}"])
        if j % 2 == 1:
            for col in "ABCDEFGH":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    total_debt_row = data_start2 + 4  # 22
    ws[f"A{total_debt_row}"] = "Total Debt"
    ws[f"A{total_debt_row}"].font = Font(name=FONT_NAME, bold=True)
    ws[f"B{total_debt_row}"] = "=SUM('Debt Tracker'!D4:D7)"
    style_currency(ws[f"B{total_debt_row}"])
    ws[f"B{total_debt_row}"].font = Font(name=FONT_NAME, bold=True)
    ws.merge_cells(f"C{total_debt_row}:D{total_debt_row}")
    overall_debt_pct = ("IFERROR((SUM('Debt Tracker'!C4:C7)-SUM('Debt Tracker'!D4:D7))"
                         "/SUM('Debt Tracker'!C4:C7),0)")
    ws[f"C{total_debt_row}"] = progress_bar_expr(overall_debt_pct)
    ws[f"C{total_debt_row}"].font = Font(name=FONT_NAME, size=10, bold=True, color=POSITIVE_TEXT)
    for col in "ABCD":
        apply_border(ws[f"{col}{total_debt_row}"])
        ws[f"{col}{total_debt_row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    ws.row_dimensions[23].height = 8  # spacer

    # --- Recent Expenses ---
    recent_header_row = 24
    ws.merge_cells(f"A{recent_header_row}:H{recent_header_row}")
    style_section_header(ws[f"A{recent_header_row}"], "\U0001F9FE RECENT EXPENSES")

    sub_header_row = recent_header_row + 1  # 25
    style_col_header(ws[f"A{sub_header_row}"], "Date")
    ws.merge_cells(f"B{sub_header_row}:C{sub_header_row}")
    style_col_header(ws[f"B{sub_header_row}"], "Merchant")
    ws.merge_cells(f"D{sub_header_row}:E{sub_header_row}")
    style_col_header(ws[f"D{sub_header_row}"], "Category")
    style_col_header(ws[f"F{sub_header_row}"], "Need/Want")
    ws.merge_cells(f"G{sub_header_row}:H{sub_header_row}")
    style_col_header(ws[f"G{sub_header_row}"], "Amount")

    # Hidden helper: matched row index for the 5 most recent transactions
    for k in range(5):
        ws[f"O{4 + k}"] = (f"=IFERROR(MATCH(LARGE('Expense Tracker'!$A$4:$A$503,{k + 1}),"
                           f"'Expense Tracker'!$A$4:$A$503,0),0)")

    for k in range(5):
        row = sub_header_row + 1 + k
        oref = f"O{4 + k}"
        ws[f"A{row}"] = f"=IF({oref}=0,\"\",INDEX('Expense Tracker'!$A$4:$A$503,{oref}))"
        style_date_cell(ws[f"A{row}"])
        ws.merge_cells(f"B{row}:C{row}")
        ws[f"B{row}"] = f"=IF({oref}=0,\"\",INDEX('Expense Tracker'!$B$4:$B$503,{oref}))"
        ws.merge_cells(f"D{row}:E{row}")
        ws[f"D{row}"] = f"=IF({oref}=0,\"\",INDEX('Expense Tracker'!$C$4:$C$503,{oref}))"
        ws[f"F{row}"] = f"=IF({oref}=0,\"\",INDEX('Expense Tracker'!$G$4:$G$503,{oref}))"
        ws.merge_cells(f"G{row}:H{row}")
        ws[f"G{row}"] = f"=IF({oref}=0,\"\",INDEX('Expense Tracker'!$E$4:$E$503,{oref}))"
        style_currency(ws[f"G{row}"])

        for col in "ABCDEFGH":
            apply_border(ws[f"{col}{row}"])
        if k % 2 == 1:
            for col in "ABCDEFGH":
                ws[f"{col}{row}"].fill = PatternFill("solid", fgColor=PALE_GREEN)

    for col in ("J", "K", "L", "M", "N", "O"):
        ws.column_dimensions[col].hidden = True

    ws.freeze_panes = "A3"
    finalize_sheet(ws)
    return ws


# ============================================================
# SECTION 4: MAIN
# ============================================================

def main():
    wb = Workbook()
    wb.remove(wb.active)

    build_lists(wb)
    build_start_here(wb)
    build_monthly_budget(wb)
    build_weekly_budget(wb)
    build_paycheck_budget(wb)
    build_biweekly_budget(wb)
    build_expense_tracker(wb)
    build_income_tracker(wb)
    build_savings_goals(wb)
    build_debt_tracker(wb)
    build_sinking_funds(wb)
    build_subscriptions(wb)
    build_bills_calendar(wb)
    build_annual_overview(wb)
    build_dashboard(wb)

    # Reorder tabs so Dashboard is the landing page (right after the hidden LISTS sheet)
    sheet_order = [
        "LISTS", "Dashboard", "START HERE", "Monthly Budget", "Weekly Budget",
        "Paycheck Budget", "Biweekly Budget", "Expense Tracker", "Income Tracker",
        "Savings Goals", "Debt Tracker", "Sinking Funds", "Subscriptions",
        "Bills Calendar", "Annual Overview",
    ]
    wb._sheets = [wb[name] for name in sheet_order]
    wb.active = wb.sheetnames.index("Dashboard")

    wb.properties.title = "Ultimate Budget Planner"
    wb.properties.subject = "Personal Finance Management"
    wb.properties.creator = "Budget Planner Pro"
    wb.properties.description = "Complete personal finance management system"
    wb.properties.keywords = "budget, finance, planner, tracker, savings, debt, expenses"
    wb.properties.category = "Personal Finance Template"

    out_dir = os.path.join(os.path.dirname(os.path.abspath(__file__)), "output")
    os.makedirs(out_dir, exist_ok=True)
    out_path = os.path.join(out_dir, "Ultimate Budget Planner.xlsx")

    try:
        wb.save(out_path)
    except Exception as exc:
        print("❌ Error while saving:", exc)
        raise

    print("\U0001F389 Budget Planner saved as 'Ultimate Budget Planner.xlsx'")
    print("\U0001F4C2 Location:", os.path.abspath(out_path))
    print("✅ Total sheets created:", len(wb.sheetnames))
    print("\U0001F4CA Ready to upload to Google Sheets!")


if __name__ == "__main__":
    main()
